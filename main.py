import os
import re
import json
import shutil
import asyncio
import time
from aiogram import Bot, Dispatcher, types
from aiogram.utils import executor
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, InputMediaPhoto, BotCommand, BotCommandScopeAllPrivateChats, BotCommandScopeAllGroupChats
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from openpyxl import Workbook, load_workbook
from datetime import datetime
# ================= TOKEN & GROUPS =================
API_TOKEN = os.getenv("BOT_TOKEN")
if not API_TOKEN:
    raise ValueError("❌ BOT_TOKEN environment variable not set!")
def env_int(name: str, default=None):
    value = os.getenv(name)
    if value is None or value.strip() == "":
        return default
    try:
        return int(value)
    except ValueError:
        print(f"⚠️ قيمة غير صالحة في {name}: {value}")
        return default

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.getenv("DATA_DIR", BASE_DIR)
EXCEL_TEMPLATE_PATH = os.getenv("EXCEL_TEMPLATE_PATH", "")
OLD_DATA_DIR = os.getenv("OLD_DATA_DIR", "")
OLD_STATE_FILE = os.getenv("OLD_STATE_FILE", "")
IMPORT_OLD_ON_START = os.getenv("IMPORT_OLD_ON_START", "0").strip().lower() in {"1", "true", "yes"}
IMPORT_OLD_MODE = os.getenv("IMPORT_OLD_MODE", "new").strip().lower()

ORDERS_FILE = os.path.join(DATA_DIR, "orders.xlsx")
READY_FILE = os.path.join(DATA_DIR, "orders_ready_current.xlsx")
STATE_FILE = os.path.join(DATA_DIR, "orders_state.json")

def ensure_data_dir():
    os.makedirs(DATA_DIR, exist_ok=True)

def migrate_legacy_files():
    """نسخ ملفات البيانات من المسار القديم إلى DATA_DIR إذا لزم"""
    try:
        ensure_data_dir()
        if os.path.abspath(DATA_DIR) == os.path.abspath(BASE_DIR):
            return
        legacy_files = ["orders.xlsx", "orders_state.json", "orders_ready_current.xlsx"]
        for name in legacy_files:
            src = os.path.join(BASE_DIR, name)
            dst = os.path.join(DATA_DIR, name)
            if os.path.exists(src) and not os.path.exists(dst):
                shutil.copyfile(src, dst)
                print(f"✅ تم نقل ملف قديم إلى DATA_DIR: {name}")
    except Exception as e:
        print(f"⚠️ تعذر نقل الملفات القديمة: {e}")

GROUP_NEW = env_int("GROUP_NEW_ID", -1003959320855)
GROUP_DESIGN = env_int("GROUP_DESIGN_ID", -1003959320855)
GROUP_READY = env_int("GROUP_READY_ID", -1003959320855)
GROUP_SENT = env_int("GROUP_SENT_ID", -1003959320855)
GROUP_ISSUES = env_int("GROUP_ISSUES_ID", -1003959320855)

# Optional hybrid mode:
# الطلبات الجديدة فقط داخل Topics بكروب واحد (افتراضيا نفس GROUP_NEW).
# FORUM_GROUP_ID (اختياري) + TOPIC_NEW_*_ID
FORUM_GROUP_ID = env_int("FORUM_GROUP_ID", GROUP_NEW)
TOPIC_NEW_ID = env_int("TOPIC_NEW_ID")
TOPIC_DESIGN_ID = env_int("TOPIC_DESIGN_ID", 12)
TOPIC_READY_ID = env_int("TOPIC_READY_ID", 14)
TOPIC_SENT_ID = env_int("TOPIC_SENT_ID", 16)
TOPIC_ISSUES_ID = env_int("TOPIC_ISSUES_ID", 5)

# Default topic IDs from provided topic links in GROUP_NEW
DEFAULT_TOPIC_IDS = {
    "new_printing": 3,
    "new_sport_sets": 74,
    "new_embroidery": 2,
    "new_urgent": 9,
}

CLASSIFIED_TOPIC_IDS = {
    "new_sport_sets": env_int("TOPIC_NEW_SPORT_SETS_ID") or env_int("TOPIC_SPORT_ID") or DEFAULT_TOPIC_IDS["new_sport_sets"],
    "new_embroidery": env_int("TOPIC_NEW_EMBROIDERY_ID") or env_int("TOPIC_EMBROIDERY_ID") or DEFAULT_TOPIC_IDS["new_embroidery"],
    "new_printing": env_int("TOPIC_NEW_PRINTING_ID") or env_int("TOPIC_PRINTING_ID") or DEFAULT_TOPIC_IDS["new_printing"],
    "new_urgent": env_int("TOPIC_NEW_URGENT_ID") or env_int("TOPIC_URGENT_ID") or DEFAULT_TOPIC_IDS["new_urgent"]
}

STATUS_DISPLAY_NAMES = {
    "new": "طلبات جديدة",
    "design": "تم التصميم",
    "ready": "مجهز",
    "sent": "تم الإرسال",
    "issues": "مشاكل",
    "new_sport_sets": "طلبات جديدة - سيتات رياضية",
    "new_embroidery": "طلبات جديدة - تطريز",
    "new_printing": "طلبات جديدة - طباعة",
    "new_urgent": "طلبات جديدة - مستعجل"
}

USE_FORUM_TOPICS = all(CLASSIFIED_TOPIC_IDS.get(s) for s in CLASSIFIED_TOPIC_IDS)

if USE_FORUM_TOPICS:
    new_fallback_thread = TOPIC_NEW_ID or CLASSIFIED_TOPIC_IDS["new_printing"]
    TARGETS_MAP = {
        "new": {"chat_id": FORUM_GROUP_ID, "thread_id": new_fallback_thread},
        "design": {"chat_id": FORUM_GROUP_ID, "thread_id": TOPIC_DESIGN_ID} if TOPIC_DESIGN_ID else {"chat_id": GROUP_DESIGN, "thread_id": None},
        "ready": {"chat_id": FORUM_GROUP_ID, "thread_id": TOPIC_READY_ID} if TOPIC_READY_ID else {"chat_id": GROUP_READY, "thread_id": None},
        "sent": {"chat_id": FORUM_GROUP_ID, "thread_id": TOPIC_SENT_ID} if TOPIC_SENT_ID else {"chat_id": GROUP_SENT, "thread_id": None},
        "issues": {"chat_id": FORUM_GROUP_ID, "thread_id": TOPIC_ISSUES_ID} if TOPIC_ISSUES_ID else {"chat_id": GROUP_ISSUES, "thread_id": None},
        "new_sport_sets": {"chat_id": FORUM_GROUP_ID, "thread_id": CLASSIFIED_TOPIC_IDS["new_sport_sets"] or new_fallback_thread},
        "new_embroidery": {"chat_id": FORUM_GROUP_ID, "thread_id": CLASSIFIED_TOPIC_IDS["new_embroidery"] or new_fallback_thread},
        "new_printing": {"chat_id": FORUM_GROUP_ID, "thread_id": CLASSIFIED_TOPIC_IDS["new_printing"] or new_fallback_thread},
        "new_urgent": {"chat_id": FORUM_GROUP_ID, "thread_id": CLASSIFIED_TOPIC_IDS["new_urgent"] or new_fallback_thread}
    }
    print(f"✅ وضع Topics مفعل للطلبات الجديدة داخل الكروب: {FORUM_GROUP_ID}")
else:
    TARGETS_MAP = {
        "new": {"chat_id": GROUP_NEW, "thread_id": None},
        "design": {"chat_id": GROUP_DESIGN, "thread_id": None},
        "ready": {"chat_id": GROUP_READY, "thread_id": None},
        "sent": {"chat_id": GROUP_SENT, "thread_id": None},
        "issues": {"chat_id": GROUP_ISSUES, "thread_id": None},
        "new_sport_sets": {"chat_id": GROUP_NEW, "thread_id": None},
        "new_embroidery": {"chat_id": GROUP_NEW, "thread_id": None},
        "new_printing": {"chat_id": GROUP_NEW, "thread_id": None},
        "new_urgent": {"chat_id": GROUP_NEW, "thread_id": None}
    }

def get_target(status: str) -> dict:
    return TARGETS_MAP[status]

def get_target_key(status: str):
    target = get_target(status)
    return (target["chat_id"], target["thread_id"] or 0)

def resolve_new_order_status(data: dict) -> str:
    if data.get("is_urgent"):
        return "new_urgent"

    if data.get("order_type") == "تطريز":
        return "new_embroidery"

    pieces = data.get("pieces", [])
    if len(pieces) == 1 and pieces[0] == "سيت رياضي":
        return "new_sport_sets"

    return "new_printing"

# الكليشية
FOOTER_TEXT = """
━━━━━━━━━━━━━━━━━━
🔹 يرجى التأكد من الطلب عند الاستلام.
🔹 في حال وجود خطأ أثناء الاستلام، تقدر ترجع الطلب بدون رسوم توصيل.
🔹 بعد استلام الطلب ومغادرة المندوب، أي تعديل أو نقص يتم مع رسوم توصيل جديدة.

🙏 شكراً لتفهمكم."""

bot = Bot(token=API_TOKEN)
dp = Dispatcher(bot, storage=MemoryStorage())
orders_data = {}
message_ids = {}
imported_order_ids = set()
order_id_lock = asyncio.Lock()
last_reserved_order_id = None
forwarded_media_cache = {}
forwarded_text_cache = {}
forwarded_photo_cache = {}
forwarded_last_order = {}
FORWARDED_TEXT_TTL = 90

def _encode_message_ids(ids_map: dict) -> dict:
    encoded = {}
    for order_id, targets in ids_map.items():
        encoded[str(order_id)] = {}
        for target_key, msg_list in targets.items():
            if isinstance(target_key, tuple):
                chat_id, thread_id = target_key
            else:
                # توافق مع النسخ القديمة التي كانت تستخدم chat_id فقط.
                chat_id, thread_id = int(target_key), 0
            encoded[str(order_id)][f"{chat_id}:{thread_id}"] = msg_list
    return encoded

def _decode_message_ids(ids_map: dict) -> dict:
    decoded = {}
    for order_id, targets in ids_map.items():
        oid = int(order_id)
        decoded[oid] = {}
        for key, msg_list in targets.items():
            if ":" in key:
                chat_id_str, thread_id_str = key.split(":", 1)
                decoded[oid][(int(chat_id_str), int(thread_id_str))] = msg_list
            else:
                decoded[oid][(int(key), 0)] = msg_list
    return decoded

def save_runtime_state(file_name: str = STATE_FILE):
    try:
        ensure_data_dir()
        payload = {
            "orders_data": {str(k): v for k, v in orders_data.items()},
            "message_ids": _encode_message_ids(message_ids),
            "imported_order_ids": sorted(imported_order_ids)
        }
        with open(file_name, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False)
    except Exception as e:
        print(f"⚠️ تعذر حفظ حالة البوت: {e}")

def load_runtime_state(file_name: str = STATE_FILE):
    global orders_data, message_ids, imported_order_ids
    try:
        ensure_data_dir()
        if not os.path.exists(file_name):
            return
        with open(file_name, "r", encoding="utf-8") as f:
            payload = json.load(f)
        orders_data = {int(k): v for k, v in payload.get("orders_data", {}).items()}
        message_ids = _decode_message_ids(payload.get("message_ids", {}))
        imported_order_ids = set(int(x) for x in payload.get("imported_order_ids", []))
        print(f"✅ تم تحميل حالة البوت: {len(orders_data)} طلب")
    except Exception as e:
        print(f"⚠️ تعذر تحميل حالة البوت: {e}")

# ================= ORDER STATES =================
class OrderState(StatesGroup):
    name = State()
    phone = State()
    source = State()
    city = State()
    area = State()
    urgent = State()
    urgent_note = State()
    order_type = State()
    team = State()
    team_other = State()
    sport_number = State()
    sport_weight = State()
    pieces = State()
    over_type = State()
    hand_type = State()
    shafa_color = State()
    scarf_owner = State()
    box_color = State()
    box_wood_name = State()
    dist_type = State()
    dist_count = State()
    dist_color = State()
    supplies_type = State()
    supplies_types = State()
    size = State()
    price = State()
    notes = State()
    images = State()

# ================= EDIT STATES =================
class EditOrderState(StatesGroup):
    waiting_for_field_choice = State()
    editing_field = State()

# ================= EXCEL FUNCTIONS =================
EXCEL_HEADER_ROW = [
    "ملاحظات",
    "عدد القطع\nأجباري",
    "يحتوي على ارجاع بضاعة؟",
    "هاتف المستلم\nأجباري 11 خانة",
    "تفاصيل العنوان\nأجباري",
    "شفرة المحافظة\nأجباري",
    "اسم المستلم",
    "المبلغ عراقي\nكامل بالالاف .\nفي حال عدم توفره سيعتبر 0",
    "رقم الوصل \nفي حال عدم وجود رقم وصل سيتم توليده من النظام",
    "كود الشحنة",
    "هاتف المستلم 2\n",
    "نوع البضاعة",
    "وصف البضاعة المسترجعة اوالمستبدلة"
]

CITY_CODE_MAP_DEFAULT_RAW = [
    ("بغداد", "BGD"),
    ("الناصرية ذي قار", "NAS"),
    ("ديالى", "DYL"),
    ("الكوت واسط", "KOT"),
    ("كربلاء", "KRB"),
    ("دهوك", "DOH"),
    ("بابل الحلة", "BBL"),
    ("النجف", "NJF"),
    ("البصرة", "BAS"),
    ("اربيل", "ARB"),
    ("كركوك", "KRK"),
    ("السليمانية", "SMH"),
    ("صلاح الدين", "SAH"),
    ("الانبار", "ANB"),
    ("السماوة المثنى", "SAM"),
    ("موصل", "MOS"),
    ("الموصل", "MOS"),
    ("الديوانية", "DWN"),
    ("العمارة ميسان", "AMA")
]

def init_excel_file(file_name: str = ORDERS_FILE):
    """إنشاء ملف Excel إذا كان غير موجود باستخدام نفس تنسيق العينة"""
    try:
        ensure_data_dir()
        if not os.path.exists(file_name):
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            _write_header(ws)
            wb.save(file_name)
            print(f"✅ تم إنشاء الملف: {file_name}")
        else:
            print(f"✅ الملف موجود: {file_name}")
    except Exception as e:
        print(f"❌ خطأ في إنشاء الملف: {e}")

def _normalize_city_key(name: str) -> str:
    cleaned = name.replace("-", " ")
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.strip()

CITY_CODE_MAP_DEFAULT = {
    _normalize_city_key(city): code for city, code in CITY_CODE_MAP_DEFAULT_RAW
}

CITY_CODE_MAP = {}

def _write_header(ws):
    for col_idx, value in enumerate(EXCEL_HEADER_ROW, start=1):
        ws.cell(row=1, column=col_idx, value=value)

def _write_city_codes(ws):
    # تم إيقاف كتابة مفاتيح المحافظات داخل ملف الإكسل.
    return

def load_city_code_map():
    """تحميل رموز المحافظات من القالب (أعمدة P/Q)"""
    global CITY_CODE_MAP
    CITY_CODE_MAP = dict(CITY_CODE_MAP_DEFAULT)
    try:
        if not EXCEL_TEMPLATE_PATH or not os.path.exists(EXCEL_TEMPLATE_PATH):
            return
        wb = load_workbook(EXCEL_TEMPLATE_PATH)
        ws = wb.active
        for row_idx in range(2, ws.max_row + 1):
            city = ws.cell(row=row_idx, column=16).value
            code = ws.cell(row=row_idx, column=17).value
            if city and code:
                CITY_CODE_MAP[_normalize_city_key(str(city))] = str(code)
    except Exception as e:
        print(f"⚠️ تعذر تحميل رموز المحافظات: {e}")

def get_city_code(city_name: str) -> str:
    if not city_name:
        return ""
    key = _normalize_city_key(str(city_name))
    return CITY_CODE_MAP.get(key, "")

def _find_next_order_row(ws) -> int:
    start_row = 2
    for row_idx in range(start_row, ws.max_row + 1):
        has_data = False
        for col_idx in range(1, 14):
            if ws.cell(row=row_idx, column=col_idx).value not in (None, ""):
                has_data = True
                break
        if not has_data:
            return row_idx
    return max(ws.max_row + 1, start_row)

def _coerce_price(value: str) -> int:
    normalized = normalize_price(str(value)) if value is not None else ""
    return int(normalized) if normalized.isdigit() else 0

def _excel_phone(phone: str) -> str:
    normalized = normalize_phone(phone or "")
    if normalized.startswith("0"):
        return normalized[1:]
    return normalized

def _excel_address(city: str, area: str) -> str:
    area_clean = str(area).strip() if area else ""
    if area_clean:
        return area_clean
    city_clean = str(city).strip() if city else ""
    return city_clean

def _excel_order_type(order_type: str) -> str:
    if order_type == "طباعة":
        return "طباعى"
    return order_type or ""

def _clean_optional(value: str):
    if value is None:
        return None
    cleaned = str(value).strip()
    if cleaned in {"", "لا يوجد", "لايوجد"}:
        return None
    return cleaned

def _parse_order_text(text: str) -> dict:
    if not text:
        return {}
    lines = [line.replace("*", "").strip() for line in text.splitlines()]

    def get_value(label: str) -> str:
        for line in lines:
            if label in line:
                value = line.split(label, 1)[1].replace(":", "").strip()
                return value
        return ""

    raw_order_id = get_value("طلب #")
    old_order_id = int(normalize_digits(raw_order_id)) if raw_order_id.isdigit() else None

    name = get_value("اسم الطفل") or get_value("الاسم")
    phone = get_value("الهاتف")
    source = get_value("المصدر") or "غير محدد"
    area_line = get_value("المحافظة - المنطقة")
    city = ""
    area = ""
    if area_line:
        parts = re.split(r"\s*-\s*", area_line, maxsplit=1)
        city = parts[0].strip()
        if len(parts) > 1:
            area = parts[1].strip()

    urgent_value = get_value("مستعجل")
    is_urgent = True if urgent_value.strip() == "نعم" else False

    order_type = get_value("النوع")

    team = _clean_optional(get_value("الفريق"))
    sport_number = _clean_optional(get_value("الرقم"))
    sport_weight = _clean_optional(get_value("وزن الطفل"))

    pieces_line = get_value("القطع")
    pieces = [p.strip() for p in pieces_line.split(",") if p.strip()] if pieces_line else []

    scarf_owner = _clean_optional(get_value("صاحب الوشاح"))
    shafa_color = _clean_optional(get_value("لون الشفقة"))
    over_type = _clean_optional(get_value("الأوفر"))
    hand_type = _clean_optional(get_value("الملحف"))
    box_color = _clean_optional(get_value("لون البوكس"))
    box_wood_name = _clean_optional(get_value("اسم الخشب"))
    dist_type_raw = _clean_optional(get_value("التوزيعات"))
    dist_type = dist_type_raw
    dist_types = []
    if dist_type_raw:
        parts = re.split(r"[،,]+", str(dist_type_raw))
        dist_types = [p.strip() for p in parts if p.strip()]
    dist_count = _clean_optional(get_value("عدد التوزيعات"))
    dist_color = _clean_optional(get_value("لون التوزيعات"))
    supplies_type = _clean_optional(get_value("المستلزمات"))
    size = _clean_optional(get_value("القياس"))
    price_raw = get_value("السعر")
    price = normalize_price(price_raw)

    notes = ""
    if "الملاحظات" in "\n".join(lines):
        start_idx = next((i for i, line in enumerate(lines) if "الملاحظات" in line), None)
        if start_idx is not None:
            collected = []
            for line in lines[start_idx + 1:]:
                if "━━━━━━━━" in line or "الحالة الحالية" in line:
                    break
                collected.append(line)
            notes = "\n".join([c for c in collected if c]).strip()
    notes = notes or "لا يوجد"

    return {
        "old_order_id": old_order_id,
        "name": name.strip() if name else "",
        "phone": normalize_phone(phone) if phone else "",
        "source": source,
        "city": city,
        "area": area,
        "is_urgent": is_urgent,
        "order_type": order_type,
        "team": team,
        "sport_number": sport_number,
        "sport_weight": sport_weight,
        "pieces": pieces,
        "scarf_owner": scarf_owner,
        "shafa_color": shafa_color,
        "over_type": over_type,
        "hand_type": hand_type,
        "box_color": box_color,
        "box_wood_name": box_wood_name,
        "dist_type": dist_type,
        "dist_types": dist_types,
        "dist_details": {},
        "dist_count": dist_count,
        "dist_color": dist_color,
        "supplies_type": supplies_type,
        "size": size,
        "price": price,
        "notes": notes
    }

async def _import_from_forwarded_message(msg: types.Message, text_override: str = None, images_list: list = None):
    text = text_override if text_override is not None else (msg.text or msg.caption or "")
    data = _parse_order_text(text)

    if not data.get("name") or not data.get("phone") or not data.get("pieces"):
        await msg.answer("❌ ما كدرت أقرأ الطلب. تأكد إنك ترسل رسالة الطلب الأصلية كاملة.")
        return

    order_id = await reserve_next_order_id()
    data["id"] = order_id

    old_id = data.pop("old_order_id", None)
    if old_id:
        extra_note = f"استيراد من طلب #{old_id}"
        data["notes"] = f"{data.get('notes', '').strip()}\n{extra_note}".strip()

    images_list = images_list or []
    orders_data[order_id] = {
        "data": data,
        "images": images_list,
        "current_group": resolve_new_order_status(data)
    }

    save_to_excel(data, ORDERS_FILE)
    await _post_order_to_group(order_id, data, images_list, orders_data[order_id]["current_group"])
    save_runtime_state()

    _set_forwarded_last_order(msg, order_id)

    await msg.answer(f"✅ تم استيراد الطلب كطلب جديد #{order_id}.")

def _get_old_state_path() -> str:
    if OLD_STATE_FILE:
        return OLD_STATE_FILE
    if OLD_DATA_DIR:
        return os.path.join(OLD_DATA_DIR, "orders_state.json")
    return ""

def _chunked(items: list, size: int):
    for i in range(0, len(items), size):
        yield items[i:i + size]

async def _post_images_to_group(order_id: int, images_list: list, status: str):
    if not images_list:
        return
    target = get_target(status)
    target_key = get_target_key(status)
    send_kwargs = {}
    if target["thread_id"]:
        send_kwargs["message_thread_id"] = target["thread_id"]

    for chunk in _chunked(images_list, 10):
        media = [InputMediaPhoto(media=i) for i in chunk]
        msg_group = await bot.send_media_group(chat_id=target["chat_id"], media=media, **send_kwargs)
        if order_id not in message_ids:
            message_ids[order_id] = {}
        if msg_group:
            message_ids[order_id].setdefault(target_key, [])
            message_ids[order_id][target_key].extend([m.message_id for m in msg_group])

async def _attach_images_to_order(order_id: int, images_list: list) -> bool:
    if not images_list:
        return False
    order_info = orders_data.get(order_id)
    if not order_info:
        return False
    existing = order_info.get("images", [])
    for img in images_list:
        if img not in existing:
            existing.append(img)
    order_info["images"] = existing
    await _post_images_to_group(order_id, images_list, order_info.get("current_group", "new"))
    save_runtime_state()
    return True

async def _post_order_to_group(order_id: int, data: dict, images_list: list, status: str):
    text = format_order_text(data, order_id, status)
    status_kb = get_status_buttons(order_id, status)
    target = get_target(status)
    target_key = get_target_key(status)
    send_kwargs = {}
    if target["thread_id"]:
        send_kwargs["message_thread_id"] = target["thread_id"]

    if images_list:
        media = [InputMediaPhoto(media=i) for i in images_list]
        msg_group = await bot.send_media_group(chat_id=target["chat_id"], media=media, **send_kwargs)
        if order_id not in message_ids:
            message_ids[order_id] = {}
        if msg_group:
            message_ids[order_id][target_key] = [m.message_id for m in msg_group]

    msg_text = await bot.send_message(
        chat_id=target["chat_id"],
        text=text,
        reply_markup=status_kb,
        parse_mode='Markdown',
        **send_kwargs
    )

    if order_id not in message_ids:
        message_ids[order_id] = {}
    if target_key not in message_ids[order_id]:
        message_ids[order_id][target_key] = []
    message_ids[order_id][target_key].append(msg_text.message_id)

async def import_and_repost_old_orders():
    old_state_path = _get_old_state_path()
    if not old_state_path or not os.path.exists(old_state_path):
        print("ℹ️ لا يوجد ملف قديم للاستيراد")
        return

    try:
        with open(old_state_path, "r", encoding="utf-8") as f:
            payload = json.load(f)
        old_orders = {int(k): v for k, v in payload.get("orders_data", {}).items()}
    except Exception as e:
        print(f"⚠️ تعذر قراءة الملف القديم: {e}")
        return

    if not old_orders:
        print("ℹ️ لا توجد طلبات قديمة")
        return

    imported = 0
    for order_id in sorted(old_orders.keys()):
        if order_id in imported_order_ids or order_id in orders_data:
            continue

        order_info = old_orders[order_id]
        data = order_info.get("data", {})
        images_list = order_info.get("images", [])

        if not data:
            continue

        data["id"] = order_id
        status = resolve_new_order_status(data)
        if IMPORT_OLD_MODE == "original":
            status = order_info.get("current_group", status)

        orders_data[order_id] = {
            "data": data,
            "images": images_list,
            "current_group": status
        }

        save_to_excel(data, ORDERS_FILE)
        await _post_order_to_group(order_id, data, images_list, status)
        imported_order_ids.add(order_id)
        imported += 1
        await asyncio.sleep(0.3)

    if imported:
        save_runtime_state()
    print(f"✅ تم استيراد وإعادة نشر {imported} طلب")

def _collect_existing_order_ids(file_name: str = ORDERS_FILE) -> set:
    ids = set()
    try:
        ids.update(int(k) for k in orders_data.keys())
    except Exception:
        pass
    try:
        ids.update(int(x) for x in imported_order_ids)
    except Exception:
        pass

    def _add_id(value):
        if isinstance(value, int):
            ids.add(value)
        elif isinstance(value, str) and value.strip().isdigit():
            ids.add(int(value.strip()))

    try:
        if os.path.exists(file_name):
            wb = load_workbook(file_name)
            ws = wb.active
            # دعم الملفات القديمة التي كانت تضع رقم الطلب في العمود A.
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
                _add_id(row[0].value)
            # التنسيق الحالي يضع رقم الطلب في العمود I.
            for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
                _add_id(row[0].value)
    except Exception as e:
        print(f"⚠️ تعذر قراءة أرقام الطلبات من الإكسل: {e}")

    return ids

def get_next_order_id(file_name: str = ORDERS_FILE):
    """احصل على رقم الطلب التالي بدون تكرار"""
    try:
        ids = _collect_existing_order_ids(file_name)
        return max(ids) + 1 if ids else 1
    except Exception as e:
        print(f"❌ خطأ في تحديد رقم الطلب التالي: {e}")
        return 1

async def reserve_next_order_id(file_name: str = ORDERS_FILE) -> int:
    global last_reserved_order_id
    async with order_id_lock:
        if last_reserved_order_id is None:
            ids = _collect_existing_order_ids(file_name)
            last_reserved_order_id = max(ids) if ids else 0
        last_reserved_order_id += 1
        return last_reserved_order_id

def save_to_excel(data, file_name: str = ORDERS_FILE):
    """احفظ الطلب في ملف Excel"""
    try:
        init_excel_file(file_name)
        wb = load_workbook(file_name)
        ws = wb.active

        notes_field = data.get("notes", "")

        order_row = [
            notes_field,
            6,
            "",
            _excel_phone(data.get("phone", "")),
            _excel_address(data.get("city", ""), data.get("area", "")),
            get_city_code(data.get("city")),
            data.get("name", ""),
            _coerce_price(data.get("price")),
            "",
            "",
            "",
            _excel_order_type(data.get("order_type", "")),
            ""
        ]

        target_row = _find_next_order_row(ws)
        for col_idx, value in enumerate(order_row, start=1):
            ws.cell(row=target_row, column=col_idx, value=value)
        
        wb.save(file_name)
        print(f"✅ تم حفظ الطلب #{data['id']} في {file_name}")
        
    except Exception as e:
        print(f"❌ خطأ في حفظ الإكسل: {e}")

def create_ready_orders_file():
    """إنشاء ملف بالطلبات الموجودة في كروب مجهز فقط"""
    try:
        ensure_data_dir()
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        _write_header(ws)
        
        for order_id, order_info in orders_data.items():
            if order_info.get("current_group") == "ready":
                data = order_info.get("data", {})
                notes_field = data.get("notes", "")

                order_row = [
                    notes_field,
                    6,
                    "",
                    _excel_phone(data.get("phone", "")),
                    _excel_address(data.get("city", ""), data.get("area", "")),
                    get_city_code(data.get("city")),
                    data.get("name", ""),
                    _coerce_price(data.get("price")),
                    "",
                    "",
                    "",
                    _excel_order_type(data.get("order_type", "")),
                    ""
                ]

                target_row = _find_next_order_row(ws)
                for col_idx, value in enumerate(order_row, start=1):
                    ws.cell(row=target_row, column=col_idx, value=value)
        
        wb.save(READY_FILE)
        print(f"✅ تم إنشاء ملف الطلبات الجاهزة: {READY_FILE}")
        return READY_FILE
    
    except Exception as e:
        print(f"❌ خطأ في إنشاء ملف الجاهزة: {e}")
        return None

def rebuild_excel_from_orders(file_name: str = ORDERS_FILE) -> int:
    """إعادة بناء ملف الإكسل من البيانات الحالية"""
    try:
        ensure_data_dir()
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        _write_header(ws)

        row_idx = 2
        for order_id in sorted(orders_data.keys()):
            data = orders_data[order_id].get("data", {})
            notes_field = data.get("notes", "")

            order_row = [
                notes_field,
                6,
                "",
                _excel_phone(data.get("phone", "")),
                _excel_address(data.get("city", ""), data.get("area", "")),
                get_city_code(data.get("city")),
                data.get("name", ""),
                _coerce_price(data.get("price", "")),
                "",
                "",
                "",
                _excel_order_type(data.get("order_type", "")),
                ""
            ]
            for col_idx, value in enumerate(order_row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1

        wb.save(file_name)
        print(f"✅ تم إعادة بناء ملف الإكسل: {file_name}")
        return row_idx - 2
    except Exception as e:
        print(f"❌ خطأ في إعادة بناء الإكسل: {e}")
        return 0

# ================= VALIDATION FUNCTIONS =================
PERSIAN_ARABIC_DIGITS_MAP = str.maketrans({
    "۰": "0", "۱": "1", "۲": "2", "۳": "3", "۴": "4",
    "۵": "5", "۶": "6", "۷": "7", "۸": "8", "۹": "9",
    "٠": "0", "١": "1", "٢": "2", "٣": "3", "٤": "4",
    "٥": "5", "٦": "6", "٧": "7", "٨": "8", "٩": "9"
})

def normalize_digits(text: str) -> str:
    return text.translate(PERSIAN_ARABIC_DIGITS_MAP)

def normalize_phone(phone: str) -> str:
    # دعم الأرقام الفارسية/العربية وإزالة أي فواصل أو رموز
    normalized = normalize_digits(phone.strip())
    digits_only = re.sub(r"\D", "", normalized)

    # تحويل كود الدولة العراقي إلى 0
    if digits_only.startswith("00964"):
        rest = digits_only[5:]
        digits_only = rest if rest.startswith("0") else f"0{rest}"
    elif digits_only.startswith("964"):
        rest = digits_only[3:]
        digits_only = rest if rest.startswith("0") else f"0{rest}"

    return digits_only

def validate_phone(phone: str) -> bool:
    normalized = normalize_phone(phone)
    return normalized.startswith("07") and len(normalized) == 11 and normalized.isdigit()

def normalize_price(price: str) -> str:
    normalized = normalize_digits(price.strip())
    # Keep digits only (no separators at all).
    return re.sub(r"\D", "", normalized)

def validate_price(price: str) -> bool:
    price = normalize_price(price)
    return price.isdigit() and len(price) > 0

def validate_dist_count(count: str) -> bool:
    count = count.strip()
    try:
        return int(count) > 0
    except:
        return False

def validate_sport_number(num: str) -> bool:
    # للسيت الرياضي نسمح بأي كتابة على الظهر (أرقام/حروف/رموز)
    return len(num.strip()) > 0

# ================= HELPER FUNCTIONS =================
def get_status_buttons(order_id: int, current_group: str = "new") -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    
    if current_group != "new":
        kb.insert(InlineKeyboardButton("⬅️ طلبات جديدة", callback_data=f"move_{order_id}_new"))
    
    if current_group != "design":
        kb.insert(InlineKeyboardButton("✏️ تم التصميم", callback_data=f"move_{order_id}_design"))
    
    if current_group != "ready":
        kb.insert(InlineKeyboardButton("📦 مجهز", callback_data=f"move_{order_id}_ready"))
    
    if current_group != "sent":
        kb.insert(InlineKeyboardButton("✈️ تم الإرسال", callback_data=f"move_{order_id}_sent"))
    
    if current_group != "issues":
        kb.insert(InlineKeyboardButton("⚠️ مشاكل", callback_data=f"move_{order_id}_issues"))

    if current_group != "new_urgent":
        kb.insert(InlineKeyboardButton("🛎️ مستعجل", callback_data=f"mark_urgent_{order_id}"))
    
    kb.insert(InlineKeyboardButton("📝 تعديل", callback_data=f"edit_{order_id}"))
    
    return kb

def format_order_text(data: dict, order_id: int, current_group: str = "new") -> str:
    source = data.get("source", "غير محدد")
    group_display = STATUS_DISPLAY_NAMES.get(current_group, "غير معروف")
    urgent_text = "نعم" if data.get("is_urgent") else "لا"
    urgent_note_text = data.get("urgent_note") or ""

    team = data.get("team")
    sport_number = data.get("sport_number")
    sport_weight = data.get("sport_weight")
    sport_line = ""
    if team:
        sport_line += f"\n⚽ *الفريق:* {team}"
    if sport_number:
        sport_line += f"\n🔢 *الرقم:* {sport_number}"
    if sport_weight:
        sport_line += f"\n⚖️ *وزن الطفل:* {sport_weight}"

    scarf_line = f"\n🧣 *صاحب الوشاح:* {data.get('scarf_owner')}" if data.get("scarf_owner") else ""
    shafa_line = f"\n🌈 *لون الشفقة:* {data.get('shafa_color')}" if data.get("shafa_color") else ""
    supplies_line = f"\n🧰 *المستلزمات:* {data.get('supplies_type')}" if data.get("supplies_type") else ""

    dist_types = data.get("dist_types") or ([] if not data.get("dist_type") else [data.get("dist_type")])
    dist_details = data.get("dist_details", {})
    dist_items = []
    for item in dist_types:
        if not item:
            continue
        details = dist_details.get(item, {})
        parts = []
        if details.get("dist_count"):
            parts.append(f"عدد: {details['dist_count']}")
        if details.get("dist_color"):
            parts.append(f"لون: {details['dist_color']}")
        if details.get("box_color"):
            parts.append(f"لون البوكس: {details['box_color']}")
        if details.get("box_wood_name"):
            parts.append(f"اسم الخشب: {details['box_wood_name']}")
        dist_items.append(f"{item} ({'، '.join(parts)})" if parts else item)
    dist_type_line = f"\n🎉 *التوزيعات:* {', '.join(dist_items)}" if dist_items else ""

    size_line = f"\n📏 *القياس:* {data.get('size', '')}" if data.get("size") else ""
    urgent_note_line = f"\n🛎️ *ملاحظة المستعجل:* {urgent_note_text}" if urgent_note_text else ""

    text = f"""📦 *طلب #{order_id}*

👤 *اسم الطفل:* {data.get('name', '')}
📞 *الهاتف:* {data.get('phone', '')}
📱 *المصدر:* {source}
📍 *المحافظة - المنطقة:* {data.get('city', '')} - {data.get('area', '')}
⏰ *مستعجل:* {urgent_text}{urgent_note_line}

🧵 *النوع:* {data.get('order_type', '')}
{sport_line}
👕 *القطع:* {', '.join(data.get('pieces', []))}
{scarf_line}{shafa_line}{supplies_line}{dist_type_line}{size_line}
💰 *السعر:* {data.get('price', '')} دينار عراقي

📝 *الملاحظات:*
{data.get('notes', '')}

━━━━━━━━━━━━━━━━━━
📍 *الحالة الحالية:* {group_display}

{FOOTER_TEXT}"""
    return text

# ================= KEYBOARDS =================
cities_list = [
    "بغداد",
    "الناصرية - ذي قار",
    "ديالى",
    "الكوت - واسط",
    "كربلاء",
    "دهوك",
    "بابل - الحلة",
    "النجف",
    "البصرة",
    "اربيل",
    "كركوك",
    "السليمانية",
    "صلاح الدين",
    "الانبار",
    "السماوة - المثنى",
    "الموصل",
    "الديوانية",
    "العمارة - ميسان"
]

def get_cities_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    for city in cities_list:
        kb.insert(InlineKeyboardButton(f"📍 {city}", callback_data=f"city_{city}"))
    return kb

def get_sources_kb() -> InlineKeyboardMarkup:
    """لوحة مفاتيح مصادر الطلب"""
    kb = InlineKeyboardMarkup(row_width=2)
    for source in sources_list:
        kb.insert(InlineKeyboardButton(f"📱 {source}", callback_data=f"source_{source}"))
    return kb

def get_order_type_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(
        InlineKeyboardButton("🖨 طباعة", callback_data="type_print"),
        InlineKeyboardButton("🧵 تطريز", callback_data="type_emb")
    )
    return kb

def get_urgent_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("✅ نعم", callback_data="urgent_yes"),
        InlineKeyboardButton("❌ لا", callback_data="urgent_no")
    )
    return kb

teams_list = ["برشلونة", "ريال مدريد", "العراق"]

def get_teams_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    for team in teams_list:
        kb.insert(InlineKeyboardButton(f"⚽ {team}", callback_data=f"team_{team}"))
    kb.add(InlineKeyboardButton("✍️ فريق آخر", callback_data="team_other"))
    return kb

pieces_list = [
    "سيت6",
    "سيت3",
    "سيت رياضي",
    "أوفر",
    "ملحف",
    "كلو",
    "صدرية",
    "كماط وحضينة",
    "عش",
    "شفقة",
    "كفوف",
    "جواريب",
    "وشاح",
    "وشاح استقبال",
    "ارنوب",
    "قطع خاصة",
    "مستلزمات",
    "توزيعات"
]

async def route_after_piece_selection(target_message: types.Message, state: FSMContext):
    data = await state.get_data()
    if data.get("need_over") and not data.get("over_type"):
        await target_message.answer("✨ نوع الأوفر:", reply_markup=get_over_type_kb())
        await OrderState.over_type.set()
        return
    if data.get("need_hand") and not data.get("hand_type"):
        await target_message.answer("🛏 نوع الملحف:", reply_markup=get_hand_type_kb())
        await OrderState.hand_type.set()
        return
    if data.get("need_shafa") and not data.get("shafa_color"):
        await target_message.answer("🎨 اكتب لون الشفقة:")
        await OrderState.shafa_color.set()
        return
    if data.get("need_dist"):
        dist_types = data.get("dist_types", [])
        if not dist_types:
            await target_message.answer("🎉 اختر نوع التوزيعات:", reply_markup=get_dist_type_select_kb([]))
            await OrderState.dist_type.set()
            return
        asked = await _ask_next_dist_detail(target_message, state)
        if asked:
            return
    if data.get("need_scarf") and not data.get("scarf_owner"):
        await target_message.answer("🧣 صاحب الوشاح؟", reply_markup=get_scarf_owner_kb())
        await OrderState.scarf_owner.set()
        return
    if data.get("need_supplies"):
        supplies_types = data.get("supplies_types", [])
        if not supplies_types:
            await target_message.answer("🧰 اختر نوع المستلزمات:", reply_markup=get_supplies_kb([]))
            await OrderState.supplies_types.set()
            return
    await _prompt_size_or_price(target_message, state)

async def _prompt_size_or_price(target_message: types.Message, state: FSMContext):
    data = await state.get_data()
    if data.get("need_size"):
        await target_message.answer("📏 اختر القياس:", reply_markup=get_size_kb())
        await OrderState.size.set()
    else:
        await target_message.answer("💰 اكتب سعر الطلب:")
        await OrderState.price.set()

def get_pieces_kb(selected: list) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    for p in pieces_list:
        mark = "✅" if p in selected else "☐"
        kb.insert(InlineKeyboardButton(f"{mark} {p}", callback_data=f"piece_{p}"))
    kb.add(InlineKeyboardButton("✔️ تم", callback_data="done_pieces"))
    return kb

def get_over_type_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("🎀 دانتيل", callback_data="over_دانتيل"),
        InlineKeyboardButton("🧵 طباكات", callback_data="over_طباكات"),
        InlineKeyboardButton("📄 صفح", callback_data="over_صفح"),
        InlineKeyboardButton("🎀🧵 دانتيل+طباكات", callback_data="over_دانتيل+طباكات")
    )
    return kb

def get_hand_type_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("🎀 كركرش", callback_data="hand_كركرش"),
        InlineKeyboardButton("🌸 حب رمان", callback_data="hand_حب رمان")
    )
    return kb

def get_scarf_owner_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("👦 ولادي", callback_data="scarf_ولادي"),
        InlineKeyboardButton("👧 بناتي", callback_data="scarf_بناتي")
    )
    return kb

def get_box_color_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("⚪ أبيض", callback_data="box_أبيض"),
        InlineKeyboardButton("⚫ رصاصي", callback_data="box_رصاصي"),
        InlineKeyboardButton("🩷 وردي", callback_data="box_وردي"),
        InlineKeyboardButton("🩵 سماوي", callback_data="box_سماوي")
    )
    return kb

def get_dist_type_select_kb(selected: list) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    options = [
        "بوكس ككو",
        "توزيعات شمع",
        "اسم خشب",
        "ستاند طبشور",
        "ستاند girl",
        "ستاند boy",
        "التاريخ خشب",
        "صينية DR",
        "صينية DB",
        "توزيعات DM",
        "توزيعات DVF",
        "توزيعات DVS",
        "توزيعات DTF",
        "توزيعات DTS",
        "توزيعات DK",
        "توزيعات DC",
        "توزيعات DF",
        "توزيعات DS",
        "توزيعات D3",
        "توزيعات خاصة"
    ]
    for opt in options:
        mark = "✅" if opt in selected else "☐"
        kb.insert(InlineKeyboardButton(f"{mark} {opt}", callback_data=f"dist_{opt}"))
    kb.add(InlineKeyboardButton("✔️ تم", callback_data="dist_done"))
    return kb

def get_supplies_kb(selected: list = None) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    selected = selected or []
    options = [
        "تعلاكة",
        "سيت مشوطة",
        "حاملة لهاية",
        "ممية",
        "ترمز",
        "حافظة حليب",
        "سيت عناية"
    ]
    for opt in options:
        mark = "✅" if opt in selected else "☐"
        kb.insert(InlineKeyboardButton(f"{mark} {opt}", callback_data=f"supply_{opt}"))
    kb.add(InlineKeyboardButton("✔️ تم", callback_data="supply_done"))
    return kb

DIST_BOX_TYPES = {"بوكس ككو"}
DIST_COLOR_TYPES = {"توزيعات شمع"}
DIST_WOOD_TYPES = {"اسم خشب"}
DIST_COUNT_TYPES = {
    "توزيعات DM",
    "توزيعات DVF",
    "توزيعات DVS",
    "توزيعات DTF",
    "توزيعات DTS",
    "توزيعات DK",
    "توزيعات DC",
    "توزيعات DF",
    "توزيعات DS",
    "توزيعات D3",
    "توزيعات خاصة"
}

def _dist_required_steps(dist_type: str) -> list:
    steps = []
    if dist_type in DIST_WOOD_TYPES:
        steps.append("box_wood_name")
    elif dist_type in DIST_BOX_TYPES:
        steps.extend(["box_color", "box_wood_name"])
    if dist_type in DIST_COLOR_TYPES:
        steps.extend(["dist_count", "dist_color"])
    elif dist_type in DIST_COUNT_TYPES:
        steps.append("dist_count")
    return steps

async def _ask_next_dist_detail(target_message: types.Message, state: FSMContext) -> bool:
    data = await state.get_data()
    dist_types = data.get("dist_types", [])
    dist_details = data.get("dist_details", {})
    for dist_type in dist_types:
        details = dist_details.get(dist_type, {})
        for step in _dist_required_steps(dist_type):
            if not details.get(step):
                await state.update_data(dist_active_type=dist_type, dist_active_step=step)
                if step == "box_color":
                    await target_message.answer("🎁 اختر لون البوكس:", reply_markup=get_box_color_kb())
                    await OrderState.box_color.set()
                elif step == "box_wood_name":
                    await target_message.answer("🪵 اكتب اسم الخشب:")
                    await OrderState.box_wood_name.set()
                elif step == "dist_count":
                    await target_message.answer("🎉 اكتب عدد التوزيعات:")
                    await OrderState.dist_count.set()
                elif step == "dist_color":
                    await target_message.answer("🎨 اكتب لون التوزيعات:")
                    await OrderState.dist_color.set()
                return True
    await state.update_data(dist_active_type=None, dist_active_step=None)
    return False

sizes = ["حديثي ولادة", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12"]

def get_size_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=4)
    for s in sizes:
        kb.insert(InlineKeyboardButton(s, callback_data=f"size_{s}"))
    return kb

def get_edit_options_kb(order_id: int) -> InlineKeyboardMarkup:
    """لوحة مفاتيح خيارات التعديل"""
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("👤 اسم الطفل", callback_data=f"field_name_{order_id}"),
        InlineKeyboardButton("📞 الهاتف", callback_data=f"field_phone_{order_id}"),
        InlineKeyboardButton("💰 السعر", callback_data=f"field_price_{order_id}"),
        InlineKeyboardButton("📝 ملاحظات", callback_data=f"field_notes_{order_id}"),
        InlineKeyboardButton("❌ إلغاء", callback_data=f"cancel_edit_{order_id}")
    )
    return kb

# ================= HANDLERS =================
@dp.message_handler(commands=['start'])
async def cmd_start(msg: types.Message, state: FSMContext):
    await state.finish()
    await msg.answer("👋 مرحباً!\n\n/start - الصفحة الرئيسية\n/new - إنشاء طلب جديد\n/cancel - إلغاء الطلب الحالي\n/download - تحميل ملف الطلبات الجاهزة")

@dp.message_handler(commands=['new'])
async def cmd_new(msg: types.Message, state: FSMContext):
    await state.finish()
    await msg.answer("👤 اسم الطفل:")
    await OrderState.name.set()

@dp.message_handler(commands=['cancel'], state='*')
async def cmd_cancel(msg: types.Message, state: FSMContext):
    current = await state.get_state()
    await state.finish()
    if current is None:
        await msg.answer("ℹ️ لا يوجد طلب قيد الإنشاء.")
    else:
        await msg.answer("✅ تم إلغاء الطلب الحالي. ابدأ طلب جديد مباشرة عبر /new")

@dp.message_handler(commands=['download'])
async def cmd_download(msg: types.Message):
    """تحميل ملف الطلبات الموجودة في كروب مجهز فقط"""
    
    ready_orders = {oid: info for oid, info in orders_data.items() 
                    if info.get("current_group") == "ready"}
    
    if not ready_orders:
        # fallback: إرسال ملف كل الطلبات القديمة الموجود في الإكسل
        if os.path.exists(ORDERS_FILE):
            with open(ORDERS_FILE, 'rb') as file:
                await bot.send_document(
                    chat_id=msg.from_user.id,
                    document=types.InputFile(ORDERS_FILE),
                    caption="📊 لا توجد حالات مجهز محفوظة حالياً، تم إرسال أرشيف الطلبات من الإكسل"
                )
            await msg.answer("✅ تم إرسال الأرشيف من orders.xlsx")
        else:
            await msg.answer("❌ لا توجد طلبات في كروب 'مجهز' حتى الآن!")
        return
    
    try:
        file_path = create_ready_orders_file()
        
        if file_path and os.path.exists(file_path):
            with open(file_path, 'rb') as file:
                await bot.send_document(
                    chat_id=msg.from_user.id,
                    document=types.InputFile(file_path),
                    caption=f"📊 ملف الطلبات الجاهزة\n\n📦 عدد الطلبات: {len(ready_orders)}"
                )
            await msg.answer("✅ تم إرسال الملف!")
        else:
            await msg.answer("❌ حدث خطأ في إنشاء الملف!")
    
    except Exception as e:
        print(f"❌ خطأ في التحميل: {e}")
        await msg.answer(f"❌ خطأ: {str(e)}")

@dp.message_handler(commands=['rebuild_excel'])
async def cmd_rebuild_excel(msg: types.Message):
    await msg.answer("⏳ إعادة بناء ملف الإكسل...")
    count = rebuild_excel_from_orders()
    await msg.answer(f"✅ تم إعادة بناء ملف الإكسل ({count} طلب).")

def _is_forwarded_message(msg: types.Message) -> bool:
    return any([
        msg.forward_from,
        msg.forward_from_chat,
        msg.forward_sender_name,
        msg.forward_date
    ])

def _looks_like_order_text(text: str) -> bool:
    return "طلب #" in text and "اسم الطفل" in text

def _forward_cache_key(msg: types.Message) -> tuple:
    user_id = msg.from_user.id if msg.from_user else 0
    return (msg.chat.id, user_id)

def _set_forwarded_text_cache(msg: types.Message, text: str):
    if not text:
        return
    forwarded_text_cache[_forward_cache_key(msg)] = {
        "text": text,
        "ts": time.time()
    }

def _get_forwarded_text_cache(msg: types.Message, max_age: int = FORWARDED_TEXT_TTL):
    key = _forward_cache_key(msg)
    payload = forwarded_text_cache.get(key)
    if not payload:
        return None
    if time.time() - payload.get("ts", 0) > max_age:
        forwarded_text_cache.pop(key, None)
        return None
    return payload.get("text")

def _set_forwarded_last_order(msg: types.Message, order_id: int):
    forwarded_last_order[_forward_cache_key(msg)] = {
        "order_id": order_id,
        "ts": time.time()
    }

def _get_forwarded_last_order(msg: types.Message, max_age: int = FORWARDED_TEXT_TTL):
    key = _forward_cache_key(msg)
    payload = forwarded_last_order.get(key)
    if not payload:
        return None
    if time.time() - payload.get("ts", 0) > max_age:
        forwarded_last_order.pop(key, None)
        return None
    return payload.get("order_id")

def _cache_forwarded_photo(msg: types.Message, file_id: str):
    key = _forward_cache_key(msg)
    payload = forwarded_photo_cache.setdefault(key, {"photos": [], "ts": time.time()})
    payload["photos"].append(file_id)
    payload["ts"] = time.time()

def _collect_pending_forwarded_images(msg: types.Message, max_age: int = FORWARDED_TEXT_TTL) -> list:
    key = _forward_cache_key(msg)
    images = []

    payload = forwarded_photo_cache.get(key)
    if payload:
        if time.time() - payload.get("ts", 0) <= max_age:
            images.extend(payload.get("photos", []))
        forwarded_photo_cache.pop(key, None)

    stale_group_ids = []
    for group_id, group in forwarded_media_cache.items():
        if group.get("owner_key") != key:
            continue
        if group.get("caption"):
            continue
        if time.time() - group.get("last_ts", 0) <= max_age:
            images.extend(group.get("photos", []))
        stale_group_ids.append(group_id)
    for group_id in stale_group_ids:
        forwarded_media_cache.pop(group_id, None)

    return images

@dp.message_handler(content_types=types.ContentTypes.TEXT, state=None)
async def import_forwarded_order(msg: types.Message, state: FSMContext):
    if not _is_forwarded_message(msg):
        return

    text = msg.text or ""
    if not _looks_like_order_text(text):
        return

    _set_forwarded_text_cache(msg, text)
    pending_images = _collect_pending_forwarded_images(msg)
    await _import_from_forwarded_message(msg, text_override=text, images_list=pending_images)

@dp.message_handler(content_types=types.ContentTypes.PHOTO, state=None)
async def import_forwarded_order_photo(msg: types.Message, state: FSMContext):
    if not _is_forwarded_message(msg):
        return

    caption = msg.caption or ""
    cached_text = _get_forwarded_text_cache(msg)
    if msg.media_group_id:
        group = forwarded_media_cache.setdefault(msg.media_group_id, {
            "photos": [],
            "caption": "",
            "processing": False,
            "last_ts": time.time(),
            "owner_key": _forward_cache_key(msg),
            "caption_from_cache": False
        })
        group["owner_key"] = group.get("owner_key") or _forward_cache_key(msg)
        group["photos"].append(msg.photo[-1].file_id)
        if caption and _looks_like_order_text(caption):
            group["caption"] = caption
            group["caption_from_cache"] = False
        elif cached_text and not group.get("caption"):
            group["caption"] = cached_text
            group["caption_from_cache"] = True
        group["last_ts"] = time.time()

        if group["caption"] and not group["processing"]:
            group["processing"] = True
            await asyncio.sleep(0.8)
            group = forwarded_media_cache.get(msg.media_group_id)
            if not group or not group.get("caption"):
                return
            photos = group.get("photos", [])
            caption = group.get("caption", "")
            caption_from_cache = group.get("caption_from_cache")
            del forwarded_media_cache[msg.media_group_id]
            if caption_from_cache:
                order_id = _get_forwarded_last_order(msg)
                if order_id and await _attach_images_to_order(order_id, photos):
                    return
            await _import_from_forwarded_message(msg, text_override=caption, images_list=photos)
        return

    if caption and _looks_like_order_text(caption):
        await _import_from_forwarded_message(msg, text_override=caption, images_list=[msg.photo[-1].file_id])
        return

    if cached_text:
        order_id = _get_forwarded_last_order(msg)
        if order_id and await _attach_images_to_order(order_id, [msg.photo[-1].file_id]):
            return
        await _import_from_forwarded_message(msg, text_override=cached_text, images_list=[msg.photo[-1].file_id])
        return

    _cache_forwarded_photo(msg, msg.photo[-1].file_id)

@dp.message_handler(commands=['import_old'])
async def cmd_import_old(msg: types.Message):
    await msg.answer("⏳ بدء استيراد الطلبات القديمة...")
    await import_and_repost_old_orders()
    await msg.answer("✅ تم الانتهاء من الاستيراد.")

@dp.message_handler(state=OrderState.name)
async def process_name(msg: types.Message, state: FSMContext):
    name = msg.text.strip()
    if len(name) < 2:
        await msg.answer("❌ اسم الطفل قصير جداً، حاول مرة أخرى:")
        return
    await state.update_data(name=name)
    await msg.answer("📞 رقم الهاتف:")
    await OrderState.phone.set()

@dp.message_handler(state=OrderState.phone)
async def process_phone(msg: types.Message, state: FSMContext):
    raw_phone = msg.text.strip()
    normalized_phone = normalize_phone(raw_phone)
    if not validate_phone(raw_phone):
        await msg.answer("❌ رقم الهاتف غير صحيح. يجب أن يبدأ بـ 07 ويكون 11 رقم:")
        return
    await state.update_data(phone=normalized_phone)
    await msg.answer("📱 اختر مصدر الطلب:", reply_markup=get_sources_kb())
    await OrderState.source.set()

@dp.callback_query_handler(lambda c: c.data.startswith("source_"), state=OrderState.source)
async def process_source(call: types.CallbackQuery, state: FSMContext):
    source = call.data.replace("source_", "")
    await state.update_data(source=source)
    await call.message.answer("📍 اختر المحافظة:", reply_markup=get_cities_kb())
    await OrderState.city.set()

@dp.callback_query_handler(lambda c: c.data.startswith("city_"), state=OrderState.city)
async def process_city(call: types.CallbackQuery, state: FSMContext):
    city = call.data.replace("city_", "")
    await state.update_data(city=city)
    await call.message.answer("🏘 اسم المنطقة:")
    await OrderState.area.set()

@dp.message_handler(state=OrderState.area)
async def process_area(msg: types.Message, state: FSMContext):
    area = msg.text.strip()
    if len(area) < 2:
        await msg.answer("❌ اسم المنطقة قصير جداً، حاول مرة أخرى:")
        return
    await state.update_data(area=area)
    await msg.answer("⏰ هل الطلب مستعجل؟", reply_markup=get_urgent_kb())
    await OrderState.urgent.set()

@dp.callback_query_handler(lambda c: c.data.startswith("urgent_"), state=OrderState.urgent)
async def process_urgent(call: types.CallbackQuery, state: FSMContext):
    is_urgent = call.data == "urgent_yes"
    await state.update_data(is_urgent=is_urgent)
    await call.answer()
    await call.message.answer("🧵 اختر نوع الطلب:", reply_markup=get_order_type_kb())
    await OrderState.order_type.set()

@dp.message_handler(state=OrderState.urgent)
async def process_urgent_text(msg: types.Message, state: FSMContext):
    text = msg.text.strip().lower()
    if text in ["نعم", "yes", "y"]:
        await state.update_data(is_urgent=True)
        await msg.answer("🧵 اختر نوع الطلب:", reply_markup=get_order_type_kb())
        await OrderState.order_type.set()
        return
    if text in ["لا", "no", "n"]:
        await state.update_data(is_urgent=False)
        await msg.answer("🧵 اختر نوع الطلب:", reply_markup=get_order_type_kb())
        await OrderState.order_type.set()
        return
    await msg.answer("❌ اختر نعم أو لا:", reply_markup=get_urgent_kb())


@dp.message_handler(state=OrderState.urgent_note)
async def process_urgent_note(msg: types.Message, state: FSMContext):
    text = msg.text.strip()
    if not text:
        await msg.answer("❌ اكتب ملاحظة المستعجل:")
        return

    data = await state.get_data()
    order_id = data.get("urgent_order_id")
    if not order_id:
        # حالة احتياطية في حال دخل المستخدم هنا بدون تحديد طلب.
        await state.update_data(urgent_note=text, is_urgent=True)
        await msg.answer("🧵 اختر نوع الطلب:", reply_markup=get_order_type_kb())
        await OrderState.order_type.set()
        return

    if order_id not in orders_data:
        await msg.answer("❌ لم أجد الطلب المطلوب.")
        await state.finish()
        return

    orders_data[order_id]["data"]["is_urgent"] = True
    orders_data[order_id]["data"]["urgent_note"] = text

    current_group = orders_data[order_id]["current_group"]
    if current_group == "new_urgent":
        await _refresh_order_message(order_id)
        save_runtime_state()
        await msg.answer(f"✅ تم تحديث الطلب #{order_id} كمستعجل.")
    else:
        await _move_order_to_status(order_id, "new_urgent")
        await msg.answer(f"✅ تم تحويل الطلب #{order_id} إلى مستعجل.")

    await state.finish()

@dp.callback_query_handler(lambda c: c.data.startswith("mark_urgent_"))
async def mark_order_urgent(call: types.CallbackQuery, state: FSMContext):
    try:
        parts = call.data.split("_")
        order_id = int(parts[2])
        if order_id not in orders_data:
            await call.answer("❌ لم أجد الطلب!", show_alert=True)
            return

        await state.update_data(urgent_order_id=order_id)
        await call.answer()
        await call.message.answer("📝 اكتب ملاحظة المستعجل لهذا الطلب:")
        await OrderState.urgent_note.set()
    except Exception as e:
        print(f"❌ خطأ في mark_order_urgent: {e}")
        await call.answer("❌ حدث خطأ", show_alert=True)

@dp.callback_query_handler(lambda c: c.data.startswith("type_"), state=OrderState.order_type)
async def process_order_type(call: types.CallbackQuery, state: FSMContext):
    if call.data == "type_print":
        order_type = "طباعة"
    else:
        order_type = "تطريز"

    await state.update_data(order_type=order_type)

    await call.message.edit_text("👕 اختر القطع:", reply_markup=get_pieces_kb([]))
    await state.update_data(
        pieces=[],
        team=None,
        sport_number=None,
        sport_weight=None,
        over_type=None,
        hand_type=None,
        shafa_color=None,
        scarf_owner=None,
        box_color=None,
        box_wood_name=None,
        dist_type=None,
        dist_types=[],
        dist_details={},
        dist_active_type=None,
        dist_active_step=None,
        dist_count=None,
        dist_color=None,
        supplies_type=None,
        supplies_types=[],
        size=None
    )
    await OrderState.pieces.set()

@dp.callback_query_handler(lambda c: c.data.startswith("team_"), state=OrderState.team)
async def process_team(call: types.CallbackQuery, state: FSMContext):
    team_value = call.data.replace("team_", "")
    if team_value == "other":
        await call.message.answer("✍️ اكتب اسم الفريق:")
        await OrderState.team_other.set()
        return

    await state.update_data(team=team_value)
    await call.message.answer("🔢 اكتب رقم اللاعب:")
    await OrderState.sport_number.set()

@dp.message_handler(state=OrderState.team_other)
async def process_team_other(msg: types.Message, state: FSMContext):
    team_name = msg.text.strip()
    if len(team_name) < 2:
        await msg.answer("❌ اسم الفريق قصير جداً، حاول مرة أخرى:")
        return

    await state.update_data(team=team_name)
    await msg.answer("🔢 اكتب رقم اللاعب:")
    await OrderState.sport_number.set()

@dp.message_handler(state=OrderState.sport_number)
async def process_sport_number(msg: types.Message, state: FSMContext):
    sport_number = msg.text.strip()
    if not validate_sport_number(sport_number):
        await msg.answer("❌ اكتب قيمة الظهر:")
        return

    await state.update_data(sport_number=sport_number)
    await msg.answer("⚖️ اكتب وزن الطفل:")
    await OrderState.sport_weight.set()

@dp.message_handler(state=OrderState.sport_weight)
async def process_sport_weight(msg: types.Message, state: FSMContext):
    weight = msg.text.strip()
    if len(weight) < 1:
        await msg.answer("❌ اكتب وزن الطفل:")
        return
    await state.update_data(sport_weight=weight)
    await route_after_piece_selection(msg, state)

@dp.callback_query_handler(lambda c: c.data.startswith("piece_"), state=OrderState.pieces)
async def process_pieces(call: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    selected = data.get("pieces", [])
    piece = call.data.replace("piece_", "")
    if piece in selected:
        selected.remove(piece)
    else:
        selected.append(piece)
    await state.update_data(pieces=selected)
    await call.message.edit_reply_markup(reply_markup=get_pieces_kb(selected))

@dp.callback_query_handler(lambda c: c.data == "done_pieces", state=OrderState.pieces)
async def process_done_pieces(call: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    pieces = data.get("pieces", [])
    if not pieces:
        await call.answer("❌ اختر قطعة واحدة على الأقل!", show_alert=True)
        return
    need_sport = "سيت رياضي" in pieces
    need_over = any(p in pieces for p in ["أوفر", "سيت3", "سيت6"])
    need_hand = any(p in pieces for p in ["ملحف", "سيت6"])
    need_scarf = "وشاح" in pieces
    need_shafa = "شفقة" in pieces
    need_dist = "توزيعات" in pieces
    need_supplies = "مستلزمات" in pieces
    await state.update_data(
        need_sport=need_sport,
        need_over=need_over,
        need_hand=need_hand,
        need_scarf=need_scarf,
        need_shafa=need_shafa,
        need_dist=need_dist,
        need_supplies=need_supplies,
        need_box=False,
        need_box_wood=False,
        need_dist_count=False,
        need_dist_color=False,
        need_size=need_sport or need_over,
        scarf_owner=None if not need_scarf else (data.get("scarf_owner") if data.get("scarf_owner") else None),
        shafa_color=None if not need_shafa else (data.get("shafa_color") if data.get("shafa_color") else None),
        team=None if not need_sport else (data.get("team") if data.get("team") else None),
        sport_number=None if not need_sport else (data.get("sport_number") if data.get("sport_number") else None),
        sport_weight=None if not need_sport else (data.get("sport_weight") if data.get("sport_weight") else None),
        dist_type=None,
        dist_types=[] if not need_dist else (data.get("dist_types") if data.get("dist_types") else []),
        dist_details={} if not need_dist else (data.get("dist_details") if data.get("dist_details") else {}),
        dist_active_type=None,
        dist_active_step=None,
        dist_count=None,
        dist_color=None,
        box_color=None,
        box_wood_name=None,
        supplies_type=None if not need_supplies else (data.get("supplies_type") if data.get("supplies_type") else None)
    )

    if need_sport:
        await call.message.answer("⚽ اختر الفريق:", reply_markup=get_teams_kb())
        await OrderState.team.set()
        return

    await route_after_piece_selection(call.message, state)

@dp.callback_query_handler(lambda c: c.data.startswith("over_"), state=OrderState.over_type)
async def process_over_type(call: types.CallbackQuery, state: FSMContext):
    over_choice = call.data.replace("over_", "")
    await state.update_data(over_type=over_choice)
    await route_after_piece_selection(call.message, state)

@dp.callback_query_handler(lambda c: c.data.startswith("hand_"), state=OrderState.hand_type)
async def process_hand_type(call: types.CallbackQuery, state: FSMContext):
    hand_choice = call.data.replace("hand_", "")
    await state.update_data(hand_type=hand_choice)
    await route_after_piece_selection(call.message, state)

@dp.callback_query_handler(lambda c: c.data.startswith("box_"), state=OrderState.box_color)
async def process_box_color(call: types.CallbackQuery, state: FSMContext):
    box_color = call.data.replace("box_", "")
    await state.update_data(box_color=box_color)
    data = await state.get_data()
    dist_type = data.get("dist_active_type")
    if dist_type:
        dist_details = data.get("dist_details", {})
        details = dist_details.get(dist_type, {})
        details["box_color"] = box_color
        dist_details[dist_type] = details
        await state.update_data(dist_details=dist_details)
    await route_after_piece_selection(call.message, state)

@dp.callback_query_handler(lambda c: c.data.startswith("scarf_"), state=OrderState.scarf_owner)
async def process_scarf_owner(call: types.CallbackQuery, state: FSMContext):
    scarf_owner = call.data.replace("scarf_", "")
    await state.update_data(scarf_owner=scarf_owner)
    await route_after_piece_selection(call.message, state)

@dp.message_handler(state=OrderState.shafa_color)
async def process_shafa_color(msg: types.Message, state: FSMContext):
    color = msg.text.strip()
    if len(color) < 1:
        await msg.answer("❌ اكتب لون الشفقة:")
        return
    await state.update_data(shafa_color=color)
    await route_after_piece_selection(msg, state)

@dp.callback_query_handler(lambda c: c.data.startswith("dist_"), state=OrderState.dist_type)
async def process_dist_type(call: types.CallbackQuery, state: FSMContext):
    dist_type = call.data.replace("dist_", "")
    data = await state.get_data()
    selected = data.get("dist_types", [])
    if dist_type in selected:
        selected.remove(dist_type)
    else:
        selected.append(dist_type)
    await state.update_data(dist_types=selected)
    await call.message.edit_reply_markup(reply_markup=get_dist_type_select_kb(selected))
    await call.answer()

@dp.callback_query_handler(lambda c: c.data == "dist_done", state=OrderState.dist_type)
async def process_dist_done(call: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    selected = data.get("dist_types", [])
    if not selected:
        await call.answer("❌ اختر نوع واحد على الأقل!", show_alert=True)
        return
    await call.answer()
    await route_after_piece_selection(call.message, state)

@dp.message_handler(state=OrderState.box_wood_name)
async def process_box_wood_name(msg: types.Message, state: FSMContext):
    name = msg.text.strip()
    if len(name) < 1:
        await msg.answer("❌ اكتب اسم الخشب:")
        return
    await state.update_data(box_wood_name=name)
    data = await state.get_data()
    dist_type = data.get("dist_active_type")
    if dist_type:
        dist_details = data.get("dist_details", {})
        details = dist_details.get(dist_type, {})
        details["box_wood_name"] = name
        dist_details[dist_type] = details
        await state.update_data(dist_details=dist_details)
    await route_after_piece_selection(msg, state)

@dp.message_handler(state=OrderState.dist_color)
async def process_dist_color(msg: types.Message, state: FSMContext):
    color = msg.text.strip()
    if len(color) < 1:
        await msg.answer("❌ اكتب لون التوزيعات:")
        return
    await state.update_data(dist_color=color)
    data = await state.get_data()
    dist_type = data.get("dist_active_type")
    if dist_type:
        dist_details = data.get("dist_details", {})
        details = dist_details.get(dist_type, {})
        details["dist_color"] = color
        dist_details[dist_type] = details
        await state.update_data(dist_details=dist_details)
    await route_after_piece_selection(msg, state)

@dp.callback_query_handler(lambda c: c.data.startswith("supply_"), state=OrderState.supplies_types)
async def process_supplies_type(call: types.CallbackQuery, state: FSMContext):
    supply = call.data.replace("supply_", "")
    if supply == "done":
        data = await state.get_data()
        if not data.get("supplies_types"):
            await call.answer("❌ اختر نوع واحد على الأقل!", show_alert=True)
            return
        await state.update_data(supplies_type=", ".join(data.get("supplies_types", [])))
        await call.answer()
        await route_after_piece_selection(call.message, state)
        return

    data = await state.get_data()
    selected = data.get("supplies_types", [])
    if supply in selected:
        selected.remove(supply)
    else:
        selected.append(supply)
    await state.update_data(supplies_types=selected)
    await call.message.edit_reply_markup(reply_markup=get_supplies_kb(selected))
    await call.answer()

@dp.message_handler(state=OrderState.dist_count)
async def process_dist_count(msg: types.Message, state: FSMContext):
    count = msg.text.strip()
    try:
        if int(count) <= 0:
            raise ValueError
    except:
        await msg.answer("❌ أدخل رقماً صحيحاً أكبر من 0:")
        return
    await state.update_data(dist_count=count)
    data = await state.get_data()
    dist_type = data.get("dist_active_type")
    if dist_type:
        dist_details = data.get("dist_details", {})
        details = dist_details.get(dist_type, {})
        details["dist_count"] = count
        dist_details[dist_type] = details
        await state.update_data(dist_details=dist_details)
    await route_after_piece_selection(msg, state)

@dp.callback_query_handler(lambda c: c.data.startswith("size_"), state=OrderState.size)
async def process_size(call: types.CallbackQuery, state: FSMContext):
    size = call.data.replace("size_", "")
    await state.update_data(size=size)
    await call.message.answer("💰 اكتب سعر الطلب:")
    await OrderState.price.set()

@dp.message_handler(state=OrderState.price)
async def process_price(msg: types.Message, state: FSMContext):
    raw_price = msg.text.strip()
    normalized_price = normalize_price(raw_price)
    if not validate_price(raw_price):
        await msg.answer("❌ أدخل سعراً صحيحاً:")
        return
    await state.update_data(price=normalized_price)
    await msg.answer("📝 ملاحظات؟ (اكتب 'لا' بدون):")
    await OrderState.notes.set()

@dp.message_handler(state=OrderState.notes)
async def process_notes(msg: types.Message, state: FSMContext):
    notes = "لا يوجد" if msg.text.strip().lower() in ["لا", "لايوجد"] else msg.text.strip()
    await state.update_data(notes=notes)
    await msg.answer("📸 ارسل الصور (1-8) أو اكتب 'تم':")
    await state.update_data(images=[])
    await OrderState.images.set()

@dp.message_handler(content_types=['photo'], state=OrderState.images)
async def process_photo(msg: types.Message, state: FSMContext):
    data = await state.get_data()
    images = data.get("images", [])
    if len(images) >= 8:
        await msg.answer("❌ الحد الأقصى 8 صور!")
        return
    images.append(msg.photo[-1].file_id)
    await state.update_data(images=images)
    await msg.answer(f"✅ صورة ({len(images)}/8)")

@dp.message_handler(state=OrderState.images)
async def finish_order(msg: types.Message, state: FSMContext):
    if "تم" not in msg.text.lower():
        await msg.answer("❌ اكتب 'تم' أو أرسل صورة:")
        return
    
    try:
        order_id = await reserve_next_order_id()
        data = await state.get_data()
        data.pop("dist_active_type", None)
        data.pop("dist_active_step", None)
        images_list = data.get("images", [])

        data["id"] = order_id

        orders_data[order_id] = {
            "data": data,
            "images": images_list,
            "current_group": resolve_new_order_status(data)
        }
        save_runtime_state()

        save_to_excel(data, ORDERS_FILE)

        intake_status = orders_data[order_id]["current_group"]
        text = format_order_text(data, order_id, intake_status)
        status_kb = get_status_buttons(order_id, intake_status)
        target = get_target(intake_status)
        target_key = get_target_key(intake_status)
        send_kwargs = {}
        if target["thread_id"]:
            send_kwargs["message_thread_id"] = target["thread_id"]

        if images_list:
            media = [InputMediaPhoto(media=i) for i in images_list]
            msg_group = await bot.send_media_group(chat_id=target["chat_id"], media=media, **send_kwargs)
            if order_id not in message_ids:
                message_ids[order_id] = {}
            if msg_group:
                message_ids[order_id][target_key] = [m.message_id for m in msg_group]
        
        msg_text = await bot.send_message(
            chat_id=target["chat_id"], 
            text=text, 
            reply_markup=status_kb, 
            parse_mode='Markdown',
            **send_kwargs
        )
        
        if order_id not in message_ids:
            message_ids[order_id] = {}
        if target_key not in message_ids[order_id]:
            message_ids[order_id][target_key] = []
        message_ids[order_id][target_key].append(msg_text.message_id)
        
        await msg.answer(f"✅ طلب #{order_id} تم!")
    except Exception as e:
        print(f"❌ خطأ: {e}")
        await msg.answer(f"❌ خطأ: {str(e)}")
    finally:
        await state.finish()

# ================= EDIT HANDLERS =================
@dp.callback_query_handler(lambda c: c.data.startswith("edit_"))
async def edit_order_start(call: types.CallbackQuery, state: FSMContext):
    """بدء التعديل"""
    try:
        order_id = int(call.data.split("_")[1])
        
        if order_id not in orders_data:
            await call.answer("❌ لم أجد الطلب!", show_alert=True)
            return
        
        await state.update_data(edit_order_id=order_id)
        await call.message.answer(
            f"📝 اختر ما تريد تعديله في الطلب #{order_id}:",
            reply_markup=get_edit_options_kb(order_id)
        )
        await EditOrderState.waiting_for_field_choice.set()
    except Exception as e:
        print(f"❌ خطأ في edit_order_start: {e}")
        await call.answer(f"❌ خطأ: {str(e)}", show_alert=True)

@dp.callback_query_handler(lambda c: c.data.startswith("field_"), state=EditOrderState.waiting_for_field_choice)
async def choose_field(call: types.CallbackQuery, state: FSMContext):
    """اختيار الحقل المراد تعديله"""
    try:
        parts = call.data.split("_")
        field_name = parts[1]
        order_id = int(parts[2])
        
        field_display = {
            "name": "اسم الطفل",
            "phone": "الهاتف",
            "price": "السعر",
            "notes": "الملاحظات"
        }
        
        await state.update_data(edit_order_id=order_id, edit_field=field_name)
        await call.message.answer(f"✏️ اكتب القيمة الجديدة للـ {field_display.get(field_name, field_name)}:")
        await EditOrderState.editing_field.set()
    except Exception as e:
        print(f"❌ خطأ في choose_field: {e}")
        await call.answer(f"❌ خطأ: {str(e)}", show_alert=True)

@dp.message_handler(state=EditOrderState.editing_field)
async def save_edited_field(msg: types.Message, state: FSMContext):
    """حفظ التعديل"""
    try:
        data = await state.get_data()
        order_id = data.get("edit_order_id")
        field_name = data.get("edit_field")
        new_value = msg.text.strip()
        
        print(f"📝 تعديل: order_id={order_id}, field={field_name}, value={new_value}")
        
        if order_id not in orders_data:
            await msg.answer("❌ لم أجد الطلب!")
            await state.finish()
            return
        
        # التحقق من صحة البيانات
        if field_name == "phone":
            normalized_phone = normalize_phone(new_value)
            if not validate_phone(new_value):
                await msg.answer("❌ رقم الهاتف غير صحيح. يجب أن يبدأ بـ 07 ويكون 11 رقم:")
                return
            new_value = normalized_phone
        
        if field_name == "price":
            normalized_price = normalize_price(new_value)
            if not validate_price(new_value):
                await msg.answer("❌ السعر غير صحيح! حاول مرة أخرى:")
                return
            new_value = normalized_price
        
        # تحديث البيانات
        orders_data[order_id]["data"][field_name] = new_value
        print(f"✅ تم تحديث البيانات: {field_name} = {new_value}")
        
        # تحديث الرسالة في الكروب
        current_group = orders_data[order_id]["current_group"]
        current_target = get_target(current_group)
        current_target_key = get_target_key(current_group)
        
        text = format_order_text(orders_data[order_id]["data"], order_id, current_group)
        status_kb = get_status_buttons(order_id, current_group)
        
        print(f"📤 تحديث الرسالة في الوجهة {current_target}")
        
        try:
            if order_id in message_ids and current_target_key in message_ids[order_id]:
                # حدّث آخر رسالة (رسالة النص، ليست الصور)
                for msg_id in reversed(message_ids[order_id][current_target_key]):
                    try:
                        await bot.edit_message_text(
                            chat_id=current_target["chat_id"],
                            message_id=msg_id,
                            text=text,
                            reply_markup=status_kb,
                            parse_mode='Markdown'
                        )
                        print(f"✅ تم تحديث الرسالة {msg_id}")
                        break
                    except Exception as e:
                        print(f"⚠️ خطأ في تحديث الرسالة {msg_id}: {e}")
        except Exception as e:
            print(f"⚠️ خطأ في تحديث الرسالة: {e}")
        
        # حدّث Excel
        save_to_excel(orders_data[order_id]["data"], ORDERS_FILE)
        save_runtime_state()
        
        await msg.answer(f"✅ تم تعديل الطلب #{order_id} بنجاح!")
        print(f"✅✅ انتهى التعديل بنجاح")
        await state.finish()
    except Exception as e:
        print(f"❌ خطأ في save_edited_field: {e}")
        await msg.answer(f"❌ خطأ: {str(e)}")
        await state.finish()

@dp.callback_query_handler(lambda c: c.data.startswith("cancel_edit_"), state='*')
async def cancel_edit(call: types.CallbackQuery, state: FSMContext):
    """إلغاء التعديل"""
    try:
        await state.finish()
        try:
            await call.message.edit_reply_markup(reply_markup=None)
        except Exception:
            pass
        await call.answer("❌ تم إلغاء التعديل", show_alert=False)
        await call.message.answer("✅ تم إلغاء التعديل")
    except Exception as e:
        print(f"❌ خطأ: {e}")

async def _refresh_order_message(order_id: int) -> bool:
    if order_id not in orders_data:
        return False
    current_group = orders_data[order_id]["current_group"]
    current_target = get_target(current_group)
    current_target_key = get_target_key(current_group)
    text = format_order_text(orders_data[order_id]["data"], order_id, current_group)
    status_kb = get_status_buttons(order_id, current_group)

    try:
        if order_id in message_ids and current_target_key in message_ids[order_id]:
            for msg_id in reversed(message_ids[order_id][current_target_key]):
                try:
                    await bot.edit_message_text(
                        chat_id=current_target["chat_id"],
                        message_id=msg_id,
                        text=text,
                        reply_markup=status_kb,
                        parse_mode='Markdown'
                    )
                    return True
                except Exception as e:
                    print(f"⚠️ خطأ في تحديث الرسالة {msg_id}: {e}")
    except Exception as e:
        print(f"⚠️ خطأ في تحديث الرسالة: {e}")
    return False

async def _move_order_to_status(order_id: int, destination_status: str) -> bool:
    if order_id not in orders_data:
        return False

    order_info = orders_data[order_id]
    data = order_info["data"]
    images_list = order_info["images"]
    current_group = order_info["current_group"]

    if current_group == destination_status:
        return True

    target = get_target(destination_status)
    target_key = get_target_key(destination_status)
    text = format_order_text(data, order_id, destination_status)
    status_kb = get_status_buttons(order_id, destination_status)

    current_target = get_target(current_group)
    current_target_key = get_target_key(current_group)
    target_send_kwargs = {}
    if target["thread_id"]:
        target_send_kwargs["message_thread_id"] = target["thread_id"]

    try:
        if order_id in message_ids and current_target_key in message_ids[order_id]:
            for msg_id in message_ids[order_id][current_target_key]:
                try:
                    await bot.delete_message(chat_id=current_target["chat_id"], message_id=msg_id)
                except Exception as e:
                    print(f"⚠️ خطأ في حذف الرسالة {msg_id}: {e}")
            del message_ids[order_id][current_target_key]
    except Exception as e:
        print(f"⚠️ خطأ في حذف الرسائل: {e}")

    if images_list:
        media = [InputMediaPhoto(media=i) for i in images_list]
        msg_group = await bot.send_media_group(chat_id=target["chat_id"], media=media, **target_send_kwargs)
        if order_id not in message_ids:
            message_ids[order_id] = {}
        if msg_group:
            message_ids[order_id][target_key] = [m.message_id for m in msg_group]

    msg_text = await bot.send_message(
        chat_id=target["chat_id"],
        text=text,
        reply_markup=status_kb,
        parse_mode='Markdown',
        **target_send_kwargs
    )

    if order_id not in message_ids:
        message_ids[order_id] = {}
    if target_key not in message_ids[order_id]:
        message_ids[order_id][target_key] = []
    message_ids[order_id][target_key].append(msg_text.message_id)

    orders_data[order_id]["current_group"] = destination_status
    save_runtime_state()
    return True

# ================= MOVE ORDER HANDLER =================
@dp.callback_query_handler(lambda c: c.data.startswith("move_"))
async def move_order(call: types.CallbackQuery):
    try:
        parts = call.data.split("_")
        order_id = int(parts[1])
        target_group_name = parts[2]

        if order_id not in orders_data:
            await call.answer("❌ لم أجد الطلب!", show_alert=True)
            return

        order_info = orders_data[order_id]
        data = order_info["data"]
        images_list = order_info["images"]
        current_group = order_info["current_group"]

        destination_status = target_group_name
        if target_group_name == "new":
            destination_status = resolve_new_order_status(data)

        if current_group == destination_status:
            await call.answer("🔔 موجود هنا!", show_alert=True)
            return

        target = get_target(destination_status)
        target_key = get_target_key(destination_status)
        text = format_order_text(data, order_id, destination_status)
        status_kb = get_status_buttons(order_id, destination_status)

        current_target = get_target(current_group)
        current_target_key = get_target_key(current_group)
        target_send_kwargs = {}
        if target["thread_id"]:
            target_send_kwargs["message_thread_id"] = target["thread_id"]

        # ✅ حذف جميع الرسائل من الكروب السابق
        try:
            if order_id in message_ids and current_target_key in message_ids[order_id]:
                for msg_id in message_ids[order_id][current_target_key]:
                    try:
                        await bot.delete_message(chat_id=current_target["chat_id"], message_id=msg_id)
                        print(f"✅ تم حذف الرسالة {msg_id}")
                    except Exception as e:
                        print(f"⚠️ خطأ في حذف الرسالة {msg_id}: {e}")
                
                del message_ids[order_id][current_target_key]
        except Exception as e:
            print(f"⚠️ خطأ في حذف الرسائل: {e}")

        # أرسل للكروب الجديد
        if images_list:
            media = [InputMediaPhoto(media=i) for i in images_list]
            msg_group = await bot.send_media_group(chat_id=target["chat_id"], media=media, **target_send_kwargs)
            if order_id not in message_ids:
                message_ids[order_id] = {}
            if msg_group:
                message_ids[order_id][target_key] = [m.message_id for m in msg_group]
        
        msg_text = await bot.send_message(
            chat_id=target["chat_id"], 
            text=text, 
            reply_markup=status_kb, 
            parse_mode='Markdown',
            **target_send_kwargs
        )
        
        if order_id not in message_ids:
            message_ids[order_id] = {}
        if target_key not in message_ids[order_id]:
            message_ids[order_id][target_key] = []
        message_ids[order_id][target_key].append(msg_text.message_id)
        
        orders_data[order_id]["current_group"] = destination_status
        save_runtime_state()

        target_name = STATUS_DISPLAY_NAMES.get(destination_status, destination_status)
        await call.answer(f"✅ {target_name}", show_alert=False)

    except Exception as e:
        print(f"❌ خطأ: {e}")
        await call.answer(f"❌ خطأ!", show_alert=True)

if __name__ == "__main__":
    print("🚀 البوت يعمل...")
    async def on_startup(dp: Dispatcher):
        migrate_legacy_files()
        load_city_code_map()
        init_excel_file(ORDERS_FILE)
        load_runtime_state()
        if IMPORT_OLD_ON_START:
            await import_and_repost_old_orders()
        commands = [
            BotCommand("start", "بدء البوت"),
            BotCommand("new", "طلب جديد"),
            BotCommand("cancel", "إلغاء الطلب الحالي"),
            BotCommand("download", "تحميل طلبات مجهز"),
            BotCommand("import_old", "استيراد طلبات قديمة"),
            BotCommand("rebuild_excel", "إعادة بناء ملف الإكسل")
        ]
        await bot.set_my_commands(commands)
        await bot.set_my_commands(commands, scope=BotCommandScopeAllPrivateChats())
        await bot.set_my_commands(commands, scope=BotCommandScopeAllGroupChats())

    executor.start_polling(dp, skip_updates=True, on_startup=on_startup)
