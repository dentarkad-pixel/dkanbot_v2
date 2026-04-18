import os
import re
import json
from aiogram import Bot, Dispatcher, types
from aiogram.utils import executor
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, InputMediaPhoto, BotCommand, BotCommandScopeAllPrivateChats, BotCommandScopeAllGroupChats
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from openpyxl import Workbook, load_workbook
from datetime import datetime

# ================= TOKEN & GROUPS =================
API_TOKEN = os.getenv("BOT_TOKEN")

if not API_TOKEN:
    raise ValueError("❌ BOT_TOKEN environment variable not set!")

GROUP_NEW = -1003735668749
GROUP_DESIGN = -1003867470006
GROUP_READY = -1003312397488
GROUP_SENT = -1003671523271
GROUP_ISSUES = -1003747379674

def env_int(name: str, default=None):
    value = os.getenv(name)
    if value is None or value.strip() == "":
        return default
    try:
        return int(value)
    except ValueError:
        print(f"⚠️ قيمة غير صالحة في {name}: {value}")
        return default

# Optional hybrid mode:
# الطلبات الجديدة فقط داخل Topics بكروب واحد (افتراضيا نفس GROUP_NEW).
# FORUM_GROUP_ID (اختياري) + TOPIC_NEW_*_ID
FORUM_GROUP_ID = env_int("FORUM_GROUP_ID", GROUP_NEW)
TOPIC_NEW_ID = env_int("TOPIC_NEW_ID")

# Default topic IDs from provided topic links in GROUP_NEW
DEFAULT_TOPIC_IDS = {
    "new_printing": 61,
    "new_sport_sets": 59,
    "new_embroidery": 187,
    "new_urgent": 192,
}

CLASSIFIED_TOPIC_IDS = {
    "new_sport_sets": env_int("TOPIC_NEW_SPORT_SETS_ID") or env_int("TOPIC_SPORT_ID") or DEFAULT_TOPIC_IDS["new_sport_sets"],
    "new_embroidery": env_int("TOPIC_NEW_EMBROIDERY_ID") or env_int("TOPIC_EMBROIDERY_ID") or DEFAULT_TOPIC_IDS["new_embroidery"],
    "new_printing": env_int("TOPIC_NEW_PRINTING_ID") or env_int("TOPIC_PRINTING_ID") or DEFAULT_TOPIC_IDS["new_printing"],
    "new_urgent": env_int("TOPIC_NEW_URGENT_ID") or env_int("TOPIC_URGENT_ID") or DEFAULT_TOPIC_IDS["new_urgent"]
}

STATUS_DISPLAY_NAMES = {
    "new": "طلبات جديدة",
    "design": "تم التصميم",
    "ready": "مجهز",
    "sent": "تم الإرسال",
    "issues": "مشاكل",
    "new_sport_sets": "طلبات جديدة - سيتات رياضية",
    "new_embroidery": "طلبات جديدة - تطريز",
    "new_printing": "طلبات جديدة - طباعة",
    "new_urgent": "طلبات جديدة - مستعجل"
}

USE_FORUM_TOPICS = all(CLASSIFIED_TOPIC_IDS.get(s) for s in CLASSIFIED_TOPIC_IDS)

if USE_FORUM_TOPICS:
    new_fallback_thread = TOPIC_NEW_ID or CLASSIFIED_TOPIC_IDS["new_printing"]
    TARGETS_MAP = {
        "new": {"chat_id": FORUM_GROUP_ID, "thread_id": new_fallback_thread},
        "design": {"chat_id": GROUP_DESIGN, "thread_id": None},
        "ready": {"chat_id": GROUP_READY, "thread_id": None},
        "sent": {"chat_id": GROUP_SENT, "thread_id": None},
        "issues": {"chat_id": GROUP_ISSUES, "thread_id": None},
        "new_sport_sets": {"chat_id": FORUM_GROUP_ID, "thread_id": CLASSIFIED_TOPIC_IDS["new_sport_sets"] or new_fallback_thread},
        "new_embroidery": {"chat_id": FORUM_GROUP_ID, "thread_id": CLASSIFIED_TOPIC_IDS["new_embroidery"] or new_fallback_thread},
        "new_printing": {"chat_id": FORUM_GROUP_ID, "thread_id": CLASSIFIED_TOPIC_IDS["new_printing"] or new_fallback_thread},
        "new_urgent": {"chat_id": FORUM_GROUP_ID, "thread_id": CLASSIFIED_TOPIC_IDS["new_urgent"] or new_fallback_thread}
    }
    print(f"✅ وضع Topics مفعل للطلبات الجديدة داخل الكروب: {FORUM_GROUP_ID}")
else:
    TARGETS_MAP = {
        "new": {"chat_id": GROUP_NEW, "thread_id": None},
        "design": {"chat_id": GROUP_DESIGN, "thread_id": None},
        "ready": {"chat_id": GROUP_READY, "thread_id": None},
        "sent": {"chat_id": GROUP_SENT, "thread_id": None},
        "issues": {"chat_id": GROUP_ISSUES, "thread_id": None},
        "new_sport_sets": {"chat_id": GROUP_NEW, "thread_id": None},
        "new_embroidery": {"chat_id": GROUP_NEW, "thread_id": None},
        "new_printing": {"chat_id": GROUP_NEW, "thread_id": None},
        "new_urgent": {"chat_id": GROUP_NEW, "thread_id": None}
    }

def get_target(status: str) -> dict:
    return TARGETS_MAP[status]

def get_target_key(status: str):
    target = get_target(status)
    return (target["chat_id"], target["thread_id"] or 0)

def resolve_new_order_status(data: dict) -> str:
    if data.get("is_urgent"):
        return "new_urgent"

    pieces = data.get("pieces", [])
    if len(pieces) == 1 and pieces[0] == "سيت رياضي":
        return "new_sport_sets"

    if data.get("order_type") == "تطريز":
        return "new_embroidery"

    return "new_printing"

def is_order_active_in_status(order_id: int, status: str) -> bool:
    order_info = orders_data.get(order_id, {})
    if order_info.get("current_group") != status:
        return False
    target_key = get_target_key(status)
    return bool(message_ids.get(order_id, {}).get(target_key))

# الكليشية
FOOTER_TEXT = """
━━━━━━━━━━━━━━━━━━
🔹 يرجى التأكد من الطلب عند الاستلام.
🔹 في حال وجود خطأ أثناء الاستلام، تقدر ترجع الطلب بدون رسوم توصيل.
🔹 بعد استلام الطلب ومغادرة المندوب، أي تعديل أو نقص يتم مع رسوم توصيل جديدة.

🙏 شكراً لتفهمكم."""

bot = Bot(token=API_TOKEN)
dp = Dispatcher(bot, storage=MemoryStorage())
orders_data = {}
message_ids = {}
STATE_FILE = "orders_state.json"

def _encode_message_ids(ids_map: dict) -> dict:
    encoded = {}
    for order_id, targets in ids_map.items():
        encoded[str(order_id)] = {}
        for target_key, msg_list in targets.items():
            if isinstance(target_key, tuple):
                chat_id, thread_id = target_key
            else:
                # توافق مع النسخ القديمة التي كانت تستخدم chat_id فقط.
                chat_id, thread_id = int(target_key), 0
            encoded[str(order_id)][f"{chat_id}:{thread_id}"] = msg_list
    return encoded

def _decode_message_ids(ids_map: dict) -> dict:
    decoded = {}
    for order_id, targets in ids_map.items():
        oid = int(order_id)
        decoded[oid] = {}
        for key, msg_list in targets.items():
            if ":" in key:
                chat_id_str, thread_id_str = key.split(":", 1)
                decoded[oid][(int(chat_id_str), int(thread_id_str))] = msg_list
            else:
                decoded[oid][(int(key), 0)] = msg_list
    return decoded

def save_runtime_state(file_name: str = STATE_FILE):
    try:
        payload = {
            "orders_data": {str(k): v for k, v in orders_data.items()},
            "message_ids": _encode_message_ids(message_ids)
        }
        with open(file_name, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False)
    except Exception as e:
        print(f"⚠️ تعذر حفظ حالة البوت: {e}")

def load_runtime_state(file_name: str = STATE_FILE):
    global orders_data, message_ids
    try:
        if not os.path.exists(file_name):
            return
        with open(file_name, "r", encoding="utf-8") as f:
            payload = json.load(f)
        orders_data = {int(k): v for k, v in payload.get("orders_data", {}).items()}
        message_ids = _decode_message_ids(payload.get("message_ids", {}))
        print(f"✅ تم تحميل حالة البوت: {len(orders_data)} طلب")
    except Exception as e:
        print(f"⚠️ تعذر تحميل حالة البوت: {e}")

# ================= مصادر الطلب =================
sources_list = [
    "دكان",
    "أمير وأميرة",
    "واتساب",
    "تيك توك"
]

# ================= STATES =================
class OrderState(StatesGroup):
    name = State()
    phone = State()
    source = State()
    city = State()
    area = State()
    urgent = State()
    order_type = State()
    team = State()
    team_other = State()
    sport_number = State()
    child_weight = State()
    pieces = State()
    over_type = State()
    hand_type = State()
    brother_bib_type = State()
    box_color = State()
    dist_count = State()
    wax_color = State()
    care_set_color = State()
    size = State()
    price = State()
    notes = State()
    images = State()

# ================= EDIT STATES =================
class EditOrderState(StatesGroup):
    waiting_for_field_choice = State()
    editing_field = State()

# ================= EXCEL FUNCTIONS =================
def init_excel_file(file_name: str = "orders.xlsx"):
    """إنشاء ملف Excel إذا كان غير موجود"""
    try:
        if not os.path.exists(file_name):
            wb = Workbook()
            ws = wb.active
            ws.title = "Orders"
            ws.append([
                "رقم الطلب",
                "اسم الطفل",
                "الهاتف",
                "المصدر",
                "المحافظة",
                "المنطقة",
                "النوع",
                "القطع",
                "صاحب الوشاح",
                "الأوفر",
                "الملحف",
                "لون البوكس",
                "عدد التوزيعات",
                "القياس",
                "السعر",
                "ملاحظات",
                "التاريخ"
            ])
            wb.save(file_name)
            print(f"✅ تم إنشاء الملف: {file_name}")
        else:
            print(f"✅ الملف موجود: {file_name}")
    except Exception as e:
        print(f"❌ خطأ في إنشاء الملف: {e}")

def get_next_order_id(file_name: str = "orders.xlsx"):
    """احصل على رقم الطلب التالي"""
    try:
        if not os.path.exists(file_name):
            return 1
        
        wb = load_workbook(file_name)
        ws = wb.active
        ids = []
        for row in ws.iter_rows(min_row=2, max_col=1):
            if row[0].value and isinstance(row[0].value, int):
                ids.append(row[0].value)
        
        return max(ids) + 1 if ids else 1
    except Exception as e:
        print(f"❌ خطأ في قراءة الإكسل: {e}")
        return 1

def save_to_excel(data, file_name: str = "orders.xlsx"):
    """احفظ الطلب في ملف Excel"""
    try:
        init_excel_file(file_name)
        wb = load_workbook(file_name)
        ws = wb.active
        
        ws.append([
            data.get("id"),
            data.get("name"),
            data.get("phone"),
            data.get("source", "غير محدد"),
            normalize_city_name(data.get("city")),
            data.get("area"),
            data.get("order_type"),
            ",".join(data.get("pieces", [])),
            data.get("scarf_owner", "لا يوجد"),
            data.get("over_type", "لا يوجد"),
            data.get("hand_type", "لا يوجد"),
            data.get("box_color", "لا يوجد"),
            data.get("dist_count", "لا يوجد"),
            data.get("size"),
            data.get("price"),
            data.get("notes"),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ])
        
        wb.save(file_name)
        print(f"✅ تم حفظ الطلب #{data['id']} في {file_name}")
        
    except Exception as e:
        print(f"❌ خطأ في حفظ الإكسل: {e}")

def delete_order_from_excel(order_id: int, file_name: str = "orders.xlsx") -> int:
    """حذف الطلب من ملف Excel حسب رقم الطلب وإرجاع عدد الصفوف المحذوفة"""
    try:
        if not os.path.exists(file_name):
            return 0

        wb = load_workbook(file_name)
        ws = wb.active

        deleted_count = 0
        # نحذف من الأسفل للأعلى حتى لا تتغير فهارس الصفوف أثناء الحذف.
        for row_idx in range(ws.max_row, 1, -1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value == order_id:
                ws.delete_rows(row_idx)
                deleted_count += 1

        if deleted_count > 0:
            wb.save(file_name)

        return deleted_count
    except Exception as e:
        print(f"❌ خطأ في حذف الطلب من الإكسل: {e}")
        return 0

def create_ready_orders_file():
    """إنشاء ملف بالطلبات الموجودة في كروب مجهز فقط"""
    try:
        ready_file = "orders_ready_current.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Ready Orders"
        
        ws.append([
            "رقم",
            "محافظة",
            "منطقة",
            "سعر",
            "اسم",
            "نوع القطع",
            "عدد القطع",
            "نوع الطلب"
        ])
        
        for order_id, order_info in orders_data.items():
            if is_order_active_in_status(order_id, "ready"):
                data = order_info.get("data", {})
                ws.append([
                    order_id,
                    normalize_city_name(data.get("city")),
                    data.get("area"),
                    data.get("price"),
                    data.get("name"),
                    ",".join(data.get("pieces", [])),
                    len(data.get("pieces", [])),
                    data.get("order_type")
                ])
        
        wb.save(ready_file)
        print(f"✅ تم إنشاء ملف الطلبات الجاهزة: {ready_file}")
        return ready_file
    
    except Exception as e:
        print(f"❌ خطأ في إنشاء ملف الجاهزة: {e}")
        return None

# ================= VALIDATION FUNCTIONS =================
PERSIAN_ARABIC_DIGITS_MAP = str.maketrans({
    "۰": "0", "۱": "1", "۲": "2", "۳": "3", "۴": "4",
    "۵": "5", "۶": "6", "۷": "7", "۸": "8", "۹": "9",
    "٠": "0", "١": "1", "٢": "2", "٣": "3", "٤": "4",
    "٥": "5", "٦": "6", "٧": "7", "٨": "8", "٩": "9"
})

def normalize_digits(text: str) -> str:
    return text.translate(PERSIAN_ARABIC_DIGITS_MAP)

def normalize_phone(phone: str) -> str:
    # دعم الأرقام الفارسية/العربية وإزالة أي فواصل أو رموز
    normalized = normalize_digits(phone.strip())
    digits_only = re.sub(r"\D", "", normalized)

    # تحويل كود الدولة العراقي إلى 0
    if digits_only.startswith("00964"):
        rest = digits_only[5:]
        digits_only = rest if rest.startswith("0") else f"0{rest}"
    elif digits_only.startswith("964"):
        rest = digits_only[3:]
        digits_only = rest if rest.startswith("0") else f"0{rest}"

    return digits_only

def validate_phone(phone: str) -> bool:
    normalized = normalize_phone(phone)
    return normalized.startswith("07") and len(normalized) == 11 and normalized.isdigit()

def normalize_price(price: str) -> str:
    normalized = normalize_digits(price.strip())
    # Keep digits only (no separators at all).
    return re.sub(r"\D", "", normalized)

def validate_price(price: str) -> bool:
    price = normalize_price(price)
    return price.isdigit() and len(price) > 0

def validate_dist_count(count: str) -> bool:
    count = count.strip()
    try:
        return int(count) > 0
    except:
        return False

def validate_sport_number(num: str) -> bool:
    # للسيت الرياضي نسمح بأي كتابة على الظهر (أرقام/حروف/رموز)
    return len(num.strip()) > 0

def normalize_city_name(city: str) -> str:
    city_map = {
        "بغداد": "بغداد",
        "الناصرية - ذي قار": "الناصرية ذي قار",
        "الناصرية ذي قار": "الناصرية ذي قار",
        "ديالى": "ديالى",
        "الكوت - واسط": "الكوت واسط",
        "الكوت واسط": "الكوت واسط",
        "كربلاء": "كربلاء",
        "دهوك": "دهوك",
        "بابل - الحلة": "بابل الحلة",
        "بابل الحلة": "بابل الحلة",
        "النجف": "النجف",
        "البصرة": "البصرة",
        "اربيل": "اربيل",
        "كركوك": "كركوك",
        "السليمانية": "السليمانيه",
        "السليمانيه": "السليمانيه",
        "صلاح الدين": "صلاح الدين",
        "الانبار": "الانبار",
        "السماوة - المثنى": "السماوة المثنى",
        "السماوة المثنى": "السماوة المثنى",
        "الموصل": "موصل",
        "موصل": "موصل",
        "الديوانية": "الديوانية",
        "العمارة - ميسان": "العمارة ميسان",
        "العمارة ميسان": "العمارة ميسان"
    }
    return city_map.get(city, city)

# ================= HELPER FUNCTIONS =================
def get_status_buttons(order_id: int, current_group: str = "new") -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    
    if current_group != "new":
        kb.insert(InlineKeyboardButton("⬅️ طلبات جديدة", callback_data=f"move_{order_id}_new"))
    
    if current_group != "design":
        kb.insert(InlineKeyboardButton("✏️ تم التصميم", callback_data=f"move_{order_id}_design"))
    
    if current_group != "ready":
        kb.insert(InlineKeyboardButton("📦 مجهز", callback_data=f"move_{order_id}_ready"))
    
    if current_group != "sent":
        kb.insert(InlineKeyboardButton("✈️ تم الإرسال", callback_data=f"move_{order_id}_sent"))
    
    if current_group != "issues":
        kb.insert(InlineKeyboardButton("⚠️ مشاكل", callback_data=f"move_{order_id}_issues"))
    
    kb.insert(InlineKeyboardButton("📝 تعديل", callback_data=f"edit_{order_id}"))
    
    return kb

def format_order_text(data: dict, order_id: int, current_group: str = "new") -> str:
    over = data.get("over_type", "لا يوجد")
    hand = data.get("hand_type", "لا يوجد")
    box = data.get("box_color", "لا يوجد")
    dist = data.get("dist_count", "لا يوجد")
    source = data.get("source", "غير محدد")
    group_display = STATUS_DISPLAY_NAMES.get(current_group, "غير معروف")
    urgent_text = "نعم" if data.get("is_urgent") else "لا"
    child_weight = data.get("child_weight")
    brother_bib_type = data.get("brother_bib_type")
    wax_color = data.get("wax_color")
    care_set_color = data.get("care_set_color")
    team = data.get("team")
    sport_number = data.get("sport_number")

    sport_line = ""
    if team:
        sport_line += f"\n⚽ *الفريق:* {team}"
    if sport_number:
        sport_line += f"\n🔢 *الرقم:* {sport_number}"
    if child_weight:
        sport_line += f"\n⚖️ *وزن الطفل:* {child_weight}"

    extra_line = ""
    if brother_bib_type:
        extra_line += f"\n👕 *صدرية أخ:* {brother_bib_type}"
    if wax_color:
        extra_line += f"\n🕯️ *لون توزيعات الشمع:* {wax_color}"
    if care_set_color:
        extra_line += f"\n🧴 *لون سيت العناية:* {care_set_color}"

    text = f"""📦 *طلب #{order_id}*

👤 *اسم الطفل:* {data['name']}
📞 *الهاتف:* {data['phone']}
📱 *المصدر:* {source}
📍 *المحافظة - المنطقة:* {data['city']} - {data['area']}
⏰ *مستعجل:* {urgent_text}

🧵 *النوع:* {data['order_type']}
{sport_line}
👕 *القطع:* {', '.join(data['pieces'])}
{extra_line}

👗 *الأوفر:* {over}
🛏 *الملحف:* {hand}
🎁 *لون البوكس:* {box}
🎉 *عدد التوزيعات:* {dist}

📏 *القياس:* {data['size']}
💰 *السعر:* {data['price']} دينار عراقي

📝 *الملاحظات:*
{data['notes']}

━━━━━━━━━━━━━━━━━━
📍 *الحالة الحالية:* {group_display}

{FOOTER_TEXT}"""
    return text

# ================= KEYBOARDS =================
cities_list = [
    "بغداد",
    "الناصرية ذي قار",
    "ديالى",
    "الكوت واسط",
    "كربلاء",
    "دهوك",
    "بابل الحلة",
    "النجف",
    "البصرة",
    "اربيل",
    "كركوك",
    "السليمانيه",
    "صلاح الدين",
    "الانبار",
    "السماوة المثنى",
    "موصل",
    "الديوانية",
    "العمارة ميسان"
]

def get_cities_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    for city in cities_list:
        kb.insert(InlineKeyboardButton(f"📍 {city}", callback_data=f"city_{city}"))
    return kb

def get_sources_kb() -> InlineKeyboardMarkup:
    """لوحة مفاتيح مصادر الطلب"""
    kb = InlineKeyboardMarkup(row_width=2)
    for source in sources_list:
        kb.insert(InlineKeyboardButton(f"📱 {source}", callback_data=f"source_{source}"))
    return kb

def get_order_type_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.add(
        InlineKeyboardButton("🖨 طباعة", callback_data="type_print"),
        InlineKeyboardButton("🧵 تطريز", callback_data="type_emb")
    )
    return kb

def get_urgent_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("✅ نعم", callback_data="urgent_yes"),
        InlineKeyboardButton("❌ لا", callback_data="urgent_no")
    )
    return kb

teams_list = ["برشلونة", "ريال مدريد", "العراق"]

def get_teams_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    for team in teams_list:
        kb.insert(InlineKeyboardButton(f"⚽ {team}", callback_data=f"team_{team}"))
    kb.add(InlineKeyboardButton("✍️ فريق آخر", callback_data="team_other"))
    return kb

pieces_list = [
    "سيت 6", "سيت 3", "سيت رياضي", "أوفر", "ملحف", "كلو", "صدرية",
    "كماط وحضينة", "عش", "شفقة", "كفوف", "تعلاكة", "سيت مشوطة",
    "صينية للتوزيعات", "اسم الطفل خشب", "حاملة لهاية", "ممية", "ترمز",
    "حافظة حليب", "طلب قطع خاصة", "صدرية اخ", "بوكس ككو", "توزيعات ورقية",
    "توزيعات مثلثة", "توزيعات لاصقة", "توزيعات شرنك", "توزيعات خاصة",
    "توزيعات شمع", "جواريب", "سيت عناية"
]

async def route_after_piece_selection(target_message: types.Message, state: FSMContext):
    data = await state.get_data()
    if data.get("need_sport") and not data.get("team"):
        await target_message.answer("⚽ اختر الفريق:", reply_markup=get_teams_kb())
        await OrderState.team.set()
        return
    if data.get("need_sport") and not data.get("sport_number"):
        await target_message.answer("🔢 اكتب رقم اللاعب (يسمح بأي كتابة):")
        await OrderState.sport_number.set()
        return
    if data.get("need_sport") and not data.get("child_weight"):
        await target_message.answer("⚖️ اكتب وزن الطفل:")
        await OrderState.child_weight.set()
        return
    if data.get("need_over") and not data.get("over_type"):
        await target_message.answer("✨ نوع الأوفر:", reply_markup=get_over_type_kb())
        await OrderState.over_type.set()
        return
    if data.get("need_hand") and not data.get("hand_type"):
        await target_message.answer("🛏 نوع الملحف:", reply_markup=get_hand_type_kb())
        await OrderState.hand_type.set()
        return
    if data.get("need_brother_bib") and not data.get("brother_bib_type"):
        await target_message.answer("👕 صدرية أخ - اختر النوع:", reply_markup=get_brother_bib_kb())
        await OrderState.brother_bib_type.set()
        return
    if data.get("need_box") and not data.get("box_color"):
        await target_message.answer("🎁 اختر لون البوكس:", reply_markup=get_box_color_kb())
        await OrderState.box_color.set()
        return
    if data.get("need_dist") and not data.get("dist_count"):
        await target_message.answer("🎉 اكتب عدد التوزيعات:")
        await OrderState.dist_count.set()
        return
    if data.get("need_wax_color") and not data.get("wax_color"):
        await target_message.answer("🕯️ اختر لون توزيعات الشمع:", reply_markup=get_wax_color_kb())
        await OrderState.wax_color.set()
        return
    if data.get("need_care_color") and not data.get("care_set_color"):
        await target_message.answer("🧴 اختر لون سيت العناية:", reply_markup=get_care_color_kb())
        await OrderState.care_set_color.set()
        return
    await target_message.answer("📏 اختر القياس:", reply_markup=get_size_kb())
    await OrderState.size.set()

def get_pieces_kb(selected: list) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    for p in pieces_list:
        mark = "✅" if p in selected else "☐"
        kb.insert(InlineKeyboardButton(f"{mark} {p}", callback_data=f"piece_{p}"))
    kb.add(InlineKeyboardButton("✔️ تم", callback_data="done_pieces"))
    return kb

def get_over_type_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("🧵 طباكات", callback_data="over_طباكات"),
        InlineKeyboardButton("🎀 دانتيل", callback_data="over_دانتيل"),
        InlineKeyboardButton("📄 طباكات صفح", callback_data="over_طباكات صفح"),
        InlineKeyboardButton("🎀🧵 طباكات دانتيل", callback_data="over_طباكات دانتيل")
    )
    return kb

def get_hand_type_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("🎀 كركش", callback_data="hand_كركش"),
        InlineKeyboardButton("🌸 حب الرمان", callback_data="hand_حب الرمان")
    )
    return kb

def get_brother_bib_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("👶 ولادي", callback_data="brobib_ولادي"),
        InlineKeyboardButton("👧 بناتي", callback_data="brobib_بناتي")
    )
    return kb

def get_wax_color_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("⚪ أبيض", callback_data="wax_أبيض"),
        InlineKeyboardButton("🩷 وردي", callback_data="wax_وردي"),
        InlineKeyboardButton("🩵 سمائي", callback_data="wax_سمائي")
    )
    return kb

def get_care_color_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("🩷 وردي", callback_data="care_وردي"),
        InlineKeyboardButton("🩵 سمائي", callback_data="care_سمائي")
    )
    return kb

def get_box_color_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("⚪ أبيض", callback_data="box_أبيض"),
        InlineKeyboardButton("⚫ رصاصي", callback_data="box_رصاصي"),
        InlineKeyboardButton("🩷 وردي", callback_data="box_وردي"),
        InlineKeyboardButton("🩵 سمائي", callback_data="box_سمائي")
    )
    return kb

sizes = ["حديثي ولادة", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12"]

def get_size_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=4)
    for s in sizes:
        kb.insert(InlineKeyboardButton(s, callback_data=f"size_{s}"))
    return kb

def get_edit_options_kb(order_id: int) -> InlineKeyboardMarkup:
    """لوحة مفاتيح خيارات التعديل"""
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("👤 اسم الطفل", callback_data=f"field_name_{order_id}"),
        InlineKeyboardButton("📞 الهاتف", callback_data=f"field_phone_{order_id}"),
        InlineKeyboardButton("💰 السعر", callback_data=f"field_price_{order_id}"),
        InlineKeyboardButton("📝 ملاحظات", callback_data=f"field_notes_{order_id}"),
        InlineKeyboardButton("❌ إلغاء", callback_data=f"cancel_edit_{order_id}")
    )
    return kb

# ================= HANDLERS =================
@dp.message_handler(commands=['start'])
async def cmd_start(msg: types.Message, state: FSMContext):
    await state.finish()
    await msg.answer("👋 مرحباً!\n\n/start - الصفحة الرئيسية\n/new - إنشاء طلب جديد\n/cancel - إلغاء الطلب الحالي\n/download - تحميل ملف الطلبات الجاهزة\n/delete_excel <رقم الطلب> - حذف طلب من الإكسل")

@dp.message_handler(commands=['new'])
async def cmd_new(msg: types.Message, state: FSMContext):
    await state.finish()
    await msg.answer("👤 اسم الطفل:")
    await OrderState.name.set()

@dp.message_handler(commands=['cancel'], state='*')
async def cmd_cancel(msg: types.Message, state: FSMContext):
    current = await state.get_state()
    await state.finish()
    if current is None:
        await msg.answer("ℹ️ لا يوجد طلب قيد الإنشاء.")
    else:
        await msg.answer("✅ تم إلغاء الطلب الحالي. ابدأ طلب جديد مباشرة عبر /new")

@dp.message_handler(commands=['download'])
async def cmd_download(msg: types.Message):
    """تحميل ملف الطلبات الموجودة في كروب مجهز فقط"""
    
    ready_orders = {
        oid: info for oid, info in orders_data.items()
        if is_order_active_in_status(oid, "ready")
    }
    
    if not ready_orders:
        await msg.answer("❌ لا توجد طلبات في كروب 'مجهز' حتى الآن!")
        return
    
    try:
        file_path = create_ready_orders_file()
        
        if file_path and os.path.exists(file_path):
            with open(file_path, 'rb') as file:
                await bot.send_document(
                    chat_id=msg.from_user.id,
                    document=types.InputFile(file_path),
                    caption=f"📊 ملف الطلبات الجاهزة\n\n📦 عدد الطلبات: {len(ready_orders)}"
                )
            await msg.answer("✅ تم إرسال الملف!")
        else:
            await msg.answer("❌ حدث خطأ في إنشاء الملف!")
    
    except Exception as e:
        print(f"❌ خطأ في التحميل: {e}")
        await msg.answer(f"❌ خطأ: {str(e)}")

@dp.message_handler(commands=['delete_excel'])
async def cmd_delete_excel(msg: types.Message):
    """حذف طلب محفوظ في orders.xlsx عند الحفظ بالخطأ"""
    args = msg.get_args().strip() if msg.get_args() else ""
    if not args:
        await msg.answer("ℹ️ الاستخدام: /delete_excel رقم_الطلب\nمثال: /delete_excel 15")
        return

    normalized_args = normalize_digits(args)
    if not re.fullmatch(r"\d+", normalized_args):
        await msg.answer("❌ رقم الطلب غير صالح. اكتب رقم صحيح فقط.")
        return

    order_id = int(normalized_args)
    if order_id <= 0:
        await msg.answer("❌ رقم الطلب يجب أن يكون أكبر من 0.")
        return

    deleted_count = delete_order_from_excel(order_id, "orders.xlsx")
    if deleted_count == 0:
        await msg.answer(f"ℹ️ ما لقيت طلب برقم #{order_id} داخل ملف الإكسل.")
        return

    await msg.answer(f"✅ تم حذف {deleted_count} صف من ملف الإكسل للطلب #{order_id}.")

@dp.message_handler(state=OrderState.name)
async def process_name(msg: types.Message, state: FSMContext):
    name = msg.text.strip()
    if len(name) < 2:
        await msg.answer("❌ اسم الطفل قصير جداً، حاول مرة أخرى:")
        return
    await state.update_data(name=name)
    await msg.answer("📞 رقم الهاتف:")
    await OrderState.phone.set()

@dp.message_handler(state=OrderState.phone)
async def process_phone(msg: types.Message, state: FSMContext):
    raw_phone = msg.text.strip()
    normalized_phone = normalize_phone(raw_phone)
    if not validate_phone(raw_phone):
        await msg.answer("❌ رقم الهاتف غير صحيح. يجب أن يبدأ بـ 07 ويكون 11 رقم:")
        return
    await state.update_data(phone=normalized_phone)
    await msg.answer("📱 اختر مصدر الطلب:", reply_markup=get_sources_kb())
    await OrderState.source.set()

@dp.callback_query_handler(lambda c: c.data.startswith("source_"), state=OrderState.source)
async def process_source(call: types.CallbackQuery, state: FSMContext):
    source = call.data.replace("source_", "")
    await state.update_data(source=source)
    await call.message.answer("📍 اختر المحافظة:", reply_markup=get_cities_kb())
    await OrderState.city.set()

@dp.callback_query_handler(lambda c: c.data.startswith("city_"), state=OrderState.city)
async def process_city(call: types.CallbackQuery, state: FSMContext):
    city = call.data.replace("city_", "")
    await state.update_data(city=normalize_city_name(city))
    await call.message.answer("🏘 اسم المنطقة:")
    await OrderState.area.set()

@dp.message_handler(state=OrderState.area)
async def process_area(msg: types.Message, state: FSMContext):
    area = msg.text.strip()
    if len(area) < 2:
        await msg.answer("❌ اسم المنطقة قصير جداً، حاول مرة أخرى:")
        return
    await state.update_data(area=area)
    await msg.answer("⏰ هل الطلب مستعجل؟", reply_markup=get_urgent_kb())
    await OrderState.urgent.set()

@dp.callback_query_handler(lambda c: c.data.startswith("urgent_"), state=OrderState.urgent)
async def process_urgent(call: types.CallbackQuery, state: FSMContext):
    is_urgent = call.data == "urgent_yes"
    await state.update_data(is_urgent=is_urgent)
    await call.message.answer("🧵 اختر نوع الطلب:", reply_markup=get_order_type_kb())
    await call.answer()
    await OrderState.order_type.set()

@dp.callback_query_handler(lambda c: c.data.startswith("type_"), state=OrderState.order_type)
async def process_order_type(call: types.CallbackQuery, state: FSMContext):
    if call.data == "type_print":
        order_type = "طباعة"
    else:
        order_type = "تطريز"

    await state.update_data(order_type=order_type)

    await call.message.edit_text("👕 اختر القطع:", reply_markup=get_pieces_kb([]))
    await state.update_data(
        pieces=[],
        team=None,
        sport_number=None,
        child_weight=None,
        over_type=None,
        hand_type=None,
        brother_bib_type=None,
        box_color=None,
        dist_count=None,
        wax_color=None,
        care_set_color=None
    )
    await OrderState.pieces.set()

@dp.callback_query_handler(lambda c: c.data.startswith("team_"), state=OrderState.team)
async def process_team(call: types.CallbackQuery, state: FSMContext):
    team_value = call.data.replace("team_", "")
    if team_value == "other":
        await call.message.answer("✍️ اكتب اسم الفريق:")
        await OrderState.team_other.set()
        return

    await state.update_data(team=team_value)
    await call.message.answer("🔢 اكتب رقم اللاعب (يسمح بأي كتابة):")
    await OrderState.sport_number.set()

@dp.message_handler(state=OrderState.team_other)
async def process_team_other(msg: types.Message, state: FSMContext):
    team_name = msg.text.strip()
    if len(team_name) < 2:
        await msg.answer("❌ اسم الفريق قصير جداً، حاول مرة أخرى:")
        return

    await state.update_data(team=team_name)
    await msg.answer("🔢 اكتب رقم اللاعب (يسمح بأي كتابة):")
    await OrderState.sport_number.set()

@dp.message_handler(state=OrderState.sport_number)
async def process_sport_number(msg: types.Message, state: FSMContext):
    sport_number = msg.text.strip()
    if not validate_sport_number(sport_number):
        await msg.answer("❌ اكتب قيمة الظهر:")
        return

    await state.update_data(sport_number=sport_number)
    await msg.answer("⚖️ اكتب وزن الطفل:")
    await OrderState.child_weight.set()

@dp.message_handler(state=OrderState.child_weight)
async def process_child_weight(msg: types.Message, state: FSMContext):
    child_weight = msg.text.strip()
    if len(child_weight) < 1:
        await msg.answer("❌ اكتب وزن الطفل:")
        return

    await state.update_data(child_weight=child_weight)
    await route_after_piece_selection(msg, state)

@dp.callback_query_handler(lambda c: c.data.startswith("piece_"), state=OrderState.pieces)
async def process_pieces(call: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    selected = data.get("pieces", [])
    piece = call.data.replace("piece_", "")
    if piece in selected:
        selected.remove(piece)
    else:
        selected.append(piece)
    await state.update_data(pieces=selected)
    await call.message.edit_reply_markup(reply_markup=get_pieces_kb(selected))

@dp.callback_query_handler(lambda c: c.data == "done_pieces", state=OrderState.pieces)
async def process_done_pieces(call: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    pieces = data.get("pieces", [])
    if not pieces:
        await call.answer("❌ اختر قطعة واحدة على الأقل!", show_alert=True)
        return
    distribution_pieces = [
        "توزيعات ورقية", "توزيعات مثلثة", "توزيعات لاصقة",
        "توزيعات شرنك", "توزيعات خاصة", "توزيعات شمع"
    ]
    need_sport = "سيت رياضي" in pieces
    need_over = any(p in pieces for p in ["أوفر", "سيت 3", "سيت 6"])
    need_hand = any(p in pieces for p in ["ملحف", "سيت 6"])
    need_brother_bib = "صدرية اخ" in pieces
    need_box = "بوكس ككو" in pieces
    need_dist = any(p in pieces for p in distribution_pieces)
    need_wax_color = "توزيعات شمع" in pieces
    need_care_color = "سيت عناية" in pieces
    await state.update_data(
        need_sport=need_sport,
        need_over=need_over,
        need_hand=need_hand,
        need_brother_bib=need_brother_bib,
        need_box=need_box,
        need_dist=need_dist,
        need_wax_color=need_wax_color,
        need_care_color=need_care_color,
        team=None if not need_sport else (data.get("team") if data.get("team") else None),
        sport_number=None if not need_sport else (data.get("sport_number") if data.get("sport_number") else None),
        child_weight=None if not need_sport else (data.get("child_weight") if data.get("child_weight") else None),
        over_type=None if not need_over else (data.get("over_type") if data.get("over_type") else None),
        hand_type=None if not need_hand else (data.get("hand_type") if data.get("hand_type") else None),
        brother_bib_type=None if not need_brother_bib else (data.get("brother_bib_type") if data.get("brother_bib_type") else None),
        box_color=None if not need_box else (data.get("box_color") if data.get("box_color") else None),
        dist_count=None if not need_dist else (data.get("dist_count") if data.get("dist_count") else None),
        wax_color=None if not need_wax_color else (data.get("wax_color") if data.get("wax_color") else None),
        care_set_color=None if not need_care_color else (data.get("care_set_color") if data.get("care_set_color") else None)
    )

    await route_after_piece_selection(call.message, state)

@dp.callback_query_handler(lambda c: c.data.startswith("over_"), state=OrderState.over_type)
async def process_over_type(call: types.CallbackQuery, state: FSMContext):
    over_choice = call.data.replace("over_", "")
    await state.update_data(over_type=over_choice)
    await route_after_piece_selection(call.message, state)

@dp.callback_query_handler(lambda c: c.data.startswith("hand_"), state=OrderState.hand_type)
async def process_hand_type(call: types.CallbackQuery, state: FSMContext):
    hand_choice = call.data.replace("hand_", "")
    await state.update_data(hand_type=hand_choice)
    await route_after_piece_selection(call.message, state)

@dp.callback_query_handler(lambda c: c.data.startswith("brobib_"), state=OrderState.brother_bib_type)
async def process_brother_bib_type(call: types.CallbackQuery, state: FSMContext):
    bib_type = call.data.replace("brobib_", "")
    await state.update_data(brother_bib_type=bib_type)
    await route_after_piece_selection(call.message, state)

@dp.callback_query_handler(lambda c: c.data.startswith("box_"), state=OrderState.box_color)
async def process_box_color(call: types.CallbackQuery, state: FSMContext):
    box_color = call.data.replace("box_", "")
    await state.update_data(box_color=box_color)
    await route_after_piece_selection(call.message, state)

@dp.message_handler(state=OrderState.dist_count)
async def process_dist_count(msg: types.Message, state: FSMContext):
    count = msg.text.strip()
    try:
        if int(count) <= 0:
            raise ValueError
    except:
        await msg.answer("❌ أدخل رقماً صحيحاً أكبر من 0:")
        return
    await state.update_data(dist_count=count)
    await route_after_piece_selection(msg, state)

@dp.callback_query_handler(lambda c: c.data.startswith("wax_"), state=OrderState.wax_color)
async def process_wax_color(call: types.CallbackQuery, state: FSMContext):
    wax_color = call.data.replace("wax_", "")
    await state.update_data(wax_color=wax_color)
    await route_after_piece_selection(call.message, state)

@dp.callback_query_handler(lambda c: c.data.startswith("care_"), state=OrderState.care_set_color)
async def process_care_set_color(call: types.CallbackQuery, state: FSMContext):
    care_color = call.data.replace("care_", "")
    await state.update_data(care_set_color=care_color)
    await route_after_piece_selection(call.message, state)

@dp.callback_query_handler(lambda c: c.data.startswith("size_"), state=OrderState.size)
async def process_size(call: types.CallbackQuery, state: FSMContext):
    size = call.data.replace("size_", "")
    await state.update_data(size=size)
    await call.message.answer("💰 اكتب سعر الطلب:")
    await OrderState.price.set()

@dp.message_handler(state=OrderState.price)
async def process_price(msg: types.Message, state: FSMContext):
    raw_price = msg.text.strip()
    normalized_price = normalize_price(raw_price)
    if not validate_price(raw_price):
        await msg.answer("❌ أدخل سعراً صحيحاً:")
        return
    await state.update_data(price=normalized_price)
    await msg.answer("📝 ملاحظات؟ (اكتب 'لا' بدون):")
    await OrderState.notes.set()

@dp.message_handler(state=OrderState.notes)
async def process_notes(msg: types.Message, state: FSMContext):
    notes = "لا يوجد" if msg.text.strip().lower() in ["لا", "لايوجد"] else msg.text.strip()
    await state.update_data(notes=notes)
    await msg.answer("📸 ارسل الصور (1-4) أو اكتب 'تم':")
    await state.update_data(images=[])
    await OrderState.images.set()

@dp.message_handler(content_types=['photo'], state=OrderState.images)
async def process_photo(msg: types.Message, state: FSMContext):
    data = await state.get_data()
    images = data.get("images", [])
    if len(images) >= 4:
        await msg.answer("❌ الحد الأقصى 4 صور!")
        return
    images.append(msg.photo[-1].file_id)
    await state.update_data(images=images)
    await msg.answer(f"✅ صورة ({len(images)}/4)")

@dp.message_handler(state=OrderState.images)
async def finish_order(msg: types.Message, state: FSMContext):
    if "تم" not in msg.text.lower():
        await msg.answer("❌ اكتب 'تم' أو أرسل صورة:")
        return
    
    try:
        order_id = get_next_order_id()
        data = await state.get_data()
        images_list = data.get("images", [])

        data["id"] = order_id

        orders_data[order_id] = {
            "data": data,
            "images": images_list,
            "current_group": resolve_new_order_status(data)
        }
        save_runtime_state()

        save_to_excel(data, "orders.xlsx")

        intake_status = orders_data[order_id]["current_group"]
        text = format_order_text(data, order_id, intake_status)
        status_kb = get_status_buttons(order_id, intake_status)
        target = get_target(intake_status)
        target_key = get_target_key(intake_status)
        send_kwargs = {}
        if target["thread_id"]:
            send_kwargs["message_thread_id"] = target["thread_id"]

        if images_list:
            media = [InputMediaPhoto(media=i) for i in images_list]
            msg_group = await bot.send_media_group(chat_id=target["chat_id"], media=media, **send_kwargs)
            if order_id not in message_ids:
                message_ids[order_id] = {}
            if msg_group:
                message_ids[order_id][target_key] = [m.message_id for m in msg_group]
        
        msg_text = await bot.send_message(
            chat_id=target["chat_id"], 
            text=text, 
            reply_markup=status_kb, 
            parse_mode='Markdown',
            **send_kwargs
        )
        
        if order_id not in message_ids:
            message_ids[order_id] = {}
        if target_key not in message_ids[order_id]:
            message_ids[order_id][target_key] = []
        message_ids[order_id][target_key].append(msg_text.message_id)
        save_runtime_state()
        
        await msg.answer(f"✅ طلب #{order_id} تم!")
    except Exception as e:
        print(f"❌ خطأ: {e}")
        await msg.answer(f"❌ خطأ: {str(e)}")
    finally:
        await state.finish()

# ================= EDIT HANDLERS =================
@dp.callback_query_handler(lambda c: c.data.startswith("edit_"))
async def edit_order_start(call: types.CallbackQuery, state: FSMContext):
    """بدء التعديل"""
    try:
        order_id = int(call.data.split("_")[1])
        
        if order_id not in orders_data:
            await call.answer("❌ لم أجد الطلب!", show_alert=True)
            return
        
        await state.update_data(edit_order_id=order_id)
        await call.message.answer(
            f"📝 اختر ما تريد تعديله في الطلب #{order_id}:",
            reply_markup=get_edit_options_kb(order_id)
        )
        await EditOrderState.waiting_for_field_choice.set()
    except Exception as e:
        print(f"❌ خطأ في edit_order_start: {e}")
        await call.answer(f"❌ خطأ: {str(e)}", show_alert=True)

@dp.callback_query_handler(lambda c: c.data.startswith("field_"), state=EditOrderState.waiting_for_field_choice)
async def choose_field(call: types.CallbackQuery, state: FSMContext):
    """اختيار الحقل المراد تعديله"""
    try:
        parts = call.data.split("_")
        field_name = parts[1]
        order_id = int(parts[2])
        
        field_display = {
            "name": "اسم الطفل",
            "phone": "الهاتف",
            "price": "السعر",
            "notes": "الملاحظات"
        }
        
        await state.update_data(edit_order_id=order_id, edit_field=field_name)
        await call.message.answer(f"✏️ اكتب القيمة الجديدة للـ {field_display.get(field_name, field_name)}:")
        await EditOrderState.editing_field.set()
    except Exception as e:
        print(f"❌ خطأ في choose_field: {e}")
        await call.answer(f"❌ خطأ: {str(e)}", show_alert=True)

@dp.message_handler(state=EditOrderState.editing_field)
async def save_edited_field(msg: types.Message, state: FSMContext):
    """حفظ التعديل"""
    try:
        data = await state.get_data()
        order_id = data.get("edit_order_id")
        field_name = data.get("edit_field")
        new_value = msg.text.strip()
        
        print(f"📝 تعديل: order_id={order_id}, field={field_name}, value={new_value}")
        
        if order_id not in orders_data:
            await msg.answer("❌ لم أجد الطلب!")
            await state.finish()
            return
        
        # التحقق من صحة البيانات
        if field_name == "phone":
            normalized_phone = normalize_phone(new_value)
            if not validate_phone(new_value):
                await msg.answer("❌ رقم الهاتف غير صحيح. يجب أن يبدأ بـ 07 ويكون 11 رقم:")
                return
            new_value = normalized_phone
        
        if field_name == "price":
            normalized_price = normalize_price(new_value)
            if not validate_price(new_value):
                await msg.answer("❌ السعر غير صحيح! حاول مرة أخرى:")
                return
            new_value = normalized_price
        
        # تحديث البيانات
        orders_data[order_id]["data"][field_name] = new_value
        print(f"✅ تم تحديث البيانات: {field_name} = {new_value}")
        
        # تحديث الرسالة في الكروب
        current_group = orders_data[order_id]["current_group"]
        current_target = get_target(current_group)
        current_target_key = get_target_key(current_group)
        
        text = format_order_text(orders_data[order_id]["data"], order_id, current_group)
        status_kb = get_status_buttons(order_id, current_group)
        
        print(f"📤 تحديث الرسالة في الوجهة {current_target}")
        
        try:
            if order_id in message_ids and current_target_key in message_ids[order_id]:
                # حدّث آخر رسالة (رسالة النص، ليست الصور)
                for msg_id in reversed(message_ids[order_id][current_target_key]):
                    try:
                        await bot.edit_message_text(
                            chat_id=current_target["chat_id"],
                            message_id=msg_id,
                            text=text,
                            reply_markup=status_kb,
                            parse_mode='Markdown'
                        )
                        print(f"✅ تم تحديث الرسالة {msg_id}")
                        break
                    except Exception as e:
                        print(f"⚠️ خطأ في تحديث الرسالة {msg_id}: {e}")
        except Exception as e:
            print(f"⚠️ خطأ في تحديث الرسالة: {e}")
        
        # حدّث Excel
        save_to_excel(orders_data[order_id]["data"], "orders.xlsx")
        save_runtime_state()
        
        await msg.answer(f"✅ تم تعديل الطلب #{order_id} بنجاح!")
        print(f"✅✅ انتهى التعديل بنجاح")
        await state.finish()
    except Exception as e:
        print(f"❌ خطأ في save_edited_field: {e}")
        await msg.answer(f"❌ خطأ: {str(e)}")
        await state.finish()

@dp.callback_query_handler(lambda c: c.data.startswith("cancel_edit_"), state='*')
async def cancel_edit(call: types.CallbackQuery, state: FSMContext):
    """إلغاء التعديل"""
    try:
        await state.finish()
        try:
            await call.message.edit_reply_markup(reply_markup=None)
        except Exception:
            pass
        await call.answer("❌ تم إلغاء التعديل", show_alert=False)
        await call.message.answer("✅ تم إلغاء التعديل")
    except Exception as e:
        print(f"❌ خطأ: {e}")

# ================= MOVE ORDER HANDLER =================
@dp.callback_query_handler(lambda c: c.data.startswith("move_"))
async def move_order(call: types.CallbackQuery):
    try:
        parts = call.data.split("_")
        order_id = int(parts[1])
        target_group_name = parts[2]

        if order_id not in orders_data:
            await call.answer("❌ لم أجد الطلب!", show_alert=True)
            return

        order_info = orders_data[order_id]
        data = order_info["data"]
        images_list = order_info["images"]
        current_group = order_info["current_group"]

        destination_status = target_group_name
        if target_group_name == "new":
            destination_status = resolve_new_order_status(data)

        if current_group == destination_status:
            await call.answer("🔔 موجود هنا!", show_alert=True)
            return

        target = get_target(destination_status)
        target_key = get_target_key(destination_status)
        text = format_order_text(data, order_id, destination_status)
        status_kb = get_status_buttons(order_id, destination_status)

        current_target = get_target(current_group)
        current_target_key = get_target_key(current_group)
        target_send_kwargs = {}
        if target["thread_id"]:
            target_send_kwargs["message_thread_id"] = target["thread_id"]

        # ✅ حذف جميع الرسائل من الكروب السابق
        try:
            if order_id in message_ids and current_target_key in message_ids[order_id]:
                failed_to_delete = []
                for msg_id in message_ids[order_id][current_target_key]:
                    try:
                        await bot.delete_message(chat_id=current_target["chat_id"], message_id=msg_id)
                        print(f"✅ تم حذف الرسالة {msg_id}")
                    except Exception as e:
                        print(f"⚠️ خطأ في حذف الرسالة {msg_id}: {e}")
                        failed_to_delete.append(msg_id)

                # إذا فشل الحذف (غالبا بسبب قدم الرسالة) نخليها غير قابلة للتعديل ونوضح أنها نُقلت.
                if failed_to_delete:
                    text_msg_id = message_ids[order_id][current_target_key][-1]
                    try:
                        await bot.edit_message_reply_markup(
                            chat_id=current_target["chat_id"],
                            message_id=text_msg_id,
                            reply_markup=None
                        )
                    except Exception:
                        pass
                    try:
                        await bot.edit_message_text(
                            chat_id=current_target["chat_id"],
                            message_id=text_msg_id,
                            text="ℹ️ تم نقل هذا الطلب إلى حالة أخرى.",
                            parse_mode='Markdown'
                        )
                    except Exception:
                        pass
                
                del message_ids[order_id][current_target_key]
        except Exception as e:
            print(f"⚠️ خطأ في حذف الرسائل: {e}")

        # أرسل للكروب الجديد
        if images_list:
            media = [InputMediaPhoto(media=i) for i in images_list]
            msg_group = await bot.send_media_group(chat_id=target["chat_id"], media=media, **target_send_kwargs)
            if order_id not in message_ids:
                message_ids[order_id] = {}
            if msg_group:
                message_ids[order_id][target_key] = [m.message_id for m in msg_group]
        
        msg_text = await bot.send_message(
            chat_id=target["chat_id"], 
            text=text, 
            reply_markup=status_kb, 
            parse_mode='Markdown',
            **target_send_kwargs
        )
        
        if order_id not in message_ids:
            message_ids[order_id] = {}
        if target_key not in message_ids[order_id]:
            message_ids[order_id][target_key] = []
        message_ids[order_id][target_key].append(msg_text.message_id)
        
        orders_data[order_id]["current_group"] = destination_status
        save_runtime_state()

        target_name = STATUS_DISPLAY_NAMES.get(destination_status, destination_status)
        await call.answer(f"✅ {target_name}", show_alert=False)

    except Exception as e:
        print(f"❌ خطأ: {e}")
        await call.answer(f"❌ خطأ!", show_alert=True)

if __name__ == "__main__":
    print("🚀 البوت يعمل...")
    init_excel_file("orders.xlsx")
    async def on_startup(dp: Dispatcher):
        load_runtime_state()
        commands = [
            BotCommand("start", "بدء البوت"),
            BotCommand("new", "طلب جديد"),
            BotCommand("cancel", "إلغاء الطلب الحالي"),
            BotCommand("download", "تحميل طلبات مجهز"),
            BotCommand("delete_excel", "حذف طلب من الإكسل")
        ]
        await bot.set_my_commands(commands)
        await bot.set_my_commands(commands, scope=BotCommandScopeAllPrivateChats())
        await bot.set_my_commands(commands, scope=BotCommandScopeAllGroupChats())

    executor.start_polling(dp, skip_updates=True, on_startup=on_startup)
